"""Client-facing xls generation for the pricing module (F3).

Two deliverables built from a saved price proposal (propuneri_pret):
  - listing / price-change files per client template (the layouts replicate
    the real files each retailer expects - see the source samples described
    in docs/plans/2026-07-05-modul-pricing-ofertare.md par.3);
  - a commercial offer with embedded product photos.

Template assignment is DATA (clienti_pricing.template_listare); this module
only knows how to draw each layout. Unknown/missing template -> 'generic'.
"""
import os
import re
from io import BytesIO

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from paths import BASE_DIR

IMG_DIR = os.path.join(BASE_DIR, 'app', 'static', 'product_images')
IMG_FETCH_TIMEOUT = 5     # seconds per photo download
IMG_MAX_PX = 96           # thumbnail box in the offer sheet

TITLE_FONT = Font(bold=True, size=13)
HDR_FONT = Font(bold=True)


def _autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _num(cell, value, fmt='0.00'):
    cell.value = value
    cell.number_format = fmt
    return cell


def _header_row(ws, row, labels):
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row, c, label)
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='center')


# ── listing templates ────────────────────────────────────────────────────────

def _t_kaufland(ws, data, valabil):
    ws.cell(2, 3, 'Torb Logistic - formular modificare preturi').font = TITLE_FONT
    _header_row(ws, 4, ['Cod articol', 'Cod Kaufland', 'DENUMIRE PRODUS',
                        'Pret lista vechi', 'Pret achizitie vechi',
                        'Pret lista nou', 'Pret achizitie nou', 'Valabil de la'])
    r = 5
    for li in data['linii']:
        ws.cell(r, 1, li['sku'])
        ws.cell(r, 2, li['cod_intern'])
        ws.cell(r, 3, li['descriere'])
        _num(ws.cell(r, 4), li['pret_actual'])
        _num(ws.cell(r, 5), li['pret_actual'])
        _num(ws.cell(r, 6), li['pret_propus'])
        _num(ws.cell(r, 7), li['pret_propus'])
        ws.cell(r, 8, valabil)
        r += 1
    _autosize(ws, [12, 12, 46, 13, 15, 13, 15, 13])


def _t_selgros(ws, data, valabil):
    ws.cell(1, 4, 'Lista de preturi').font = TITLE_FONT
    ws.cell(3, 1, 'Nume furnizor: TORB LOGISTIC')
    ws.cell(4, 1, f'valabil de la: {valabil}')
    _header_row(ws, 6, ['Nr. Articol\nfurnizor', 'Nr. articol\nSelgros',
                        'Denumire produs', 'UC', 'UV', 'Pret LISTA\nBAX',
                        'Pret LISTA\nBUCATA', 'Disc.\n%', 'Pret netto\nper BAX',
                        'Pret netto\nper BUCATA', 'Cod EAN\nindividual',
                        'Numar bax\npe palet'])
    r = 7
    for li in data['linii']:
        uc = li['buc_bax'] or li['buc_cutie'] or 1
        ws.cell(r, 1, li['sku'])
        ws.cell(r, 2, li['cod_intern'])
        ws.cell(r, 3, li['descriere'])
        ws.cell(r, 4, uc)
        ws.cell(r, 5, 1)
        _num(ws.cell(r, 6), round(li['pret_propus'] * uc, 2))
        _num(ws.cell(r, 7), li['pret_propus'])
        _num(ws.cell(r, 8), 0, '0.0')
        _num(ws.cell(r, 9), round(li['pret_propus'] * uc, 2))
        _num(ws.cell(r, 10), li['pret_propus'])
        ws.cell(r, 11, li['ean'])
        ws.cell(r, 12, li['bax_palet'])
        r += 1
    _autosize(ws, [12, 12, 42, 6, 6, 11, 11, 7, 11, 12, 15, 10])


def _t_fildas(ws, data, valabil):
    _header_row(ws, 3, ['Cod art fz', 'DENUMIRE PRODUS', 'GRAMAJ',
                        'PRET FACTURARE VECHI FARA TVA',
                        'PRET VANZARE NOU FARA TVA'])
    r = 5
    for li in data['linii']:
        ws.cell(r, 1, li['sku'])
        ws.cell(r, 2, li['descriere'])
        ws.cell(r, 3, f"{li['gramaj']:g}g" if li['gramaj'] else None)
        _num(ws.cell(r, 4), li['pret_actual'])
        _num(ws.cell(r, 5), li['pret_propus'])
        r += 1
    _autosize(ws, [12, 46, 10, 26, 24])


def _t_sezamo(ws, data, valabil):
    _header_row(ws, 3, ['', 'Denumire produs', 'Cod produs',
                        'Pret facturare vechi fara TVA',
                        'Pret facturare nou fara TVA'])
    r = 5
    for li in data['linii']:
        ws.cell(r, 1, li['sku'])
        ws.cell(r, 2, li['descriere'])
        ws.cell(r, 3, li['cod_intern'])
        _num(ws.cell(r, 4), li['pret_actual'])
        _num(ws.cell(r, 5), li['pret_propus'])
        r += 1
    _autosize(ws, [12, 46, 12, 26, 24])


def _t_generic(ws, data, valabil):
    ws.cell(1, 1, f"Lista de preturi TORB LOGISTIC - {data['nume_client']}")\
        .font = TITLE_FONT
    ws.cell(2, 1, f'Valabil de la: {valabil}')
    _header_row(ws, 4, ['Cod articol', 'Cod client', 'Denumire produs',
                        'Gramaj', 'EAN', 'Pret vechi fara TVA',
                        'Pret nou fara TVA'])
    r = 5
    for li in data['linii']:
        ws.cell(r, 1, li['sku'])
        ws.cell(r, 2, li['cod_intern'])
        ws.cell(r, 3, li['descriere'])
        ws.cell(r, 4, f"{li['gramaj']:g}g" if li['gramaj'] else None)
        ws.cell(r, 5, li['ean'])
        _num(ws.cell(r, 6), li['pret_actual'])
        _num(ws.cell(r, 7), li['pret_propus'])
        r += 1
    _autosize(ws, [12, 12, 46, 10, 15, 18, 18])


TEMPLATES = {
    'kaufland_modificare': _t_kaufland,
    'selgros_lista': _t_selgros,
    'fildas_lista': _t_fildas,
    'sezamo_lista': _t_sezamo,
    'generic': _t_generic,
}


def build_listare(data, template=None, valabil=''):
    """openpyxl Workbook for a proposal in the requested client layout."""
    builder = TEMPLATES.get(template or data.get('template') or 'generic',
                            _t_generic)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lista preturi'
    builder(ws, data, valabil)
    return wb


# ── commercial offer with photos ─────────────────────────────────────────────

def _safe_name(sku):
    return re.sub(r'[^A-Za-z0-9_-]', '_', str(sku))


def _local_photo(li):
    """Local photo path for a line; downloads url_sursa into the static
    cache on first use. Returns None when no usable photo."""
    if li['poza_path']:
        p = li['poza_path']
        if not os.path.isabs(p):
            p = os.path.join(BASE_DIR, p.lstrip('/'))
        return p if os.path.exists(p) else None
    url = li['poza_url']
    if not url:
        return None
    ext = os.path.splitext(url.split('?')[0])[1].lower() or '.jpg'
    if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
        ext = '.jpg'
    cached = os.path.join(IMG_DIR, _safe_name(li['sku']) + ext)
    if os.path.exists(cached):
        return cached
    try:
        import requests
        resp = requests.get(url, timeout=IMG_FETCH_TIMEOUT)
        resp.raise_for_status()
        os.makedirs(IMG_DIR, exist_ok=True)
        with open(cached, 'wb') as f:
            f.write(resp.content)
        return cached
    except Exception:
        return None


def _thumb(path):
    """Bounded-size XLImage or None if Pillow can't read the file."""
    try:
        from PIL import Image as PILImage
        with PILImage.open(path) as im:
            im.thumbnail((IMG_MAX_PX, IMG_MAX_PX))
            buf = BytesIO()
            im.convert('RGB').save(buf, format='PNG')
        buf.seek(0)
        img = XLImage(buf)
        return img
    except Exception:
        return None


def build_oferta(data, valabil=''):
    """Commercial offer workbook: photo, product data, proposed prices."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Oferta'
    ws.cell(1, 2, f"Oferta de pret TORB LOGISTIC - {data['nume_client']}")\
        .font = TITLE_FONT
    if valabil:
        ws.cell(2, 2, f'Valabil de la: {valabil}')
    _header_row(ws, 4, ['Poza', 'Cod articol', 'Denumire produs', 'Gramaj',
                        'Buc/bax', 'EAN', 'Pret fara TVA', 'TVA %',
                        'Pret cu TVA'])
    r = 5
    for li in data['linii']:
        ws.cell(r, 2, li['sku'])
        ws.cell(r, 3, li['descriere'])
        ws.cell(r, 4, f"{li['gramaj']:g}g" if li['gramaj'] else None)
        ws.cell(r, 5, li['buc_bax'] or li['buc_cutie'])
        ws.cell(r, 6, li['ean'])
        _num(ws.cell(r, 7), li['pret_propus'])
        tva = (li['tva_pct'] or 0) * 100
        _num(ws.cell(r, 8), tva, '0')
        _num(ws.cell(r, 9), round(li['pret_propus'] * (1 + tva / 100), 2))
        photo = _local_photo(li)
        img = _thumb(photo) if photo else None
        if img:
            ws.add_image(img, f'A{r}')
            ws.row_dimensions[r].height = IMG_MAX_PX * 0.78
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical='center')
        r += 1
    _autosize(ws, [14, 12, 46, 10, 8, 15, 13, 7, 12])
    return wb
