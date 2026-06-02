"""
Helper universal pentru export rapoarte în format Excel.
"""
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _format_value(val):
    """Normalize sqlite3.Row values for Excel."""
    if val is None:
        return ""
    if isinstance(val, (int, float, str, datetime)):
        return val
    return str(val)


def _write_sheet(ws, rows, headers=None):
    """rows = list of dict-like (sqlite3.Row or dict)."""
    if not rows:
        ws.cell(1, 1, "Nu există date pentru acest raport.")
        return

    if headers is None:
        headers = list(rows[0].keys())

    # Header row
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    max_widths = [len(h) for h in headers]
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            val = _format_value(row[h] if h in row.keys() else "")
            cell = ws.cell(row_idx, col_idx, val)
            cell.border = THIN_BORDER
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"
            max_widths[col_idx - 1] = max(max_widths[col_idx - 1], min(len(str(val)), 60))

    # Auto-width
    for col_idx, w in enumerate(max_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(w + 3, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def send_excel(sheets: dict, filename: str):
    """
    sheets = {'Sheet Name': [row_dict, ...]} — or {'Name': {'rows': [...], 'headers': [...]}}
    filename = output filename (e.g. 'forecast.xlsx')
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, data in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        if isinstance(data, dict):
            _write_sheet(ws, data.get("rows", []), data.get("headers"))
        else:
            _write_sheet(ws, data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def timestamped_filename(base: str) -> str:
    return f"{base}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


def export_comenzi_intern(comanda_id: int):
    """Format intern Torb: detaliu per SKU cu split RO/HU."""
    from io import BytesIO
    from db import query, query_one
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    header_fill = PatternFill('solid', fgColor='1E3A8A')
    header_font = Font(color='FFFFFF', bold=True)

    cmd = query_one("SELECT * FROM comenzi_furnizori WHERE id=?", (comanda_id,))
    if not cmd:
        raise ValueError(f"Comanda {comanda_id} nu există")

    linii = query("""
        SELECT l.*, COALESCE(p.descriere, l.descriere, l.sku) AS denumire
        FROM comenzi_furnizori_linii l
        LEFT JOIN produse p ON p.sku = l.sku
        WHERE l.comanda_id = ?
        ORDER BY l.sku
    """, (comanda_id,))

    moneda_row = query_one(
        "SELECT moneda FROM termene_aprovizionare WHERE furnizor=?",
        (cmd['furnizor'],)
    )
    moneda = (moneda_row or {}).get('moneda', 'EUR')

    wb = openpyxl.Workbook()

    # Sheet 1: Detaliu
    ws1 = wb.active
    ws1.title = 'Detaliu comandă'
    headers = ['Cod Mare', 'SKU', 'Cod Furnizor', 'Denumire', 'Gamă',
               'Cant. RO', 'Cant. HU', 'Total', f'Preț ({moneda})', f'Total ({moneda})']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    total_ro = total_hu = total_val = 0.0
    for row_i, line in enumerate(linii, 2):
        cant_ro  = line.get('cantitate_ro', 0) or 0
        cant_hu  = line.get('cantitate_export', 0) or 0
        total    = cant_ro + cant_hu
        pret     = line.get('pret_valuta', 0) or 0
        tot_val  = total * pret
        total_ro  += cant_ro
        total_hu  += cant_hu
        total_val += tot_val
        for col, val in enumerate([
            line.get('cod_mare', ''), line['sku'], line.get('cod_furnizor', ''),
            line['denumire'], line.get('gama', ''),
            cant_ro, cant_hu, total, pret, round(tot_val, 2),
        ], 1):
            ws1.cell(row=row_i, column=col, value=val)

    fr = len(linii) + 2
    ws1.cell(row=fr, column=6, value=total_ro).font = Font(bold=True)
    ws1.cell(row=fr, column=7, value=total_hu).font = Font(bold=True)
    ws1.cell(row=fr, column=8, value=total_ro + total_hu).font = Font(bold=True)
    ws1.cell(row=fr, column=10, value=round(total_val, 2)).font = Font(bold=True)

    # Sheet 2: Sumar piată
    ws2 = wb.create_sheet('Sumar piață')
    for col, h in enumerate(['Piață', 'Nr. SKU', 'Cantitate', f'Valoare ({moneda})'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    val_ro = round(sum((line.get('cantitate_ro', 0) or 0) * (line.get('pret_valuta', 0) or 0) for line in linii), 2)
    val_hu = round(sum((line.get('cantitate_export', 0) or 0) * (line.get('pret_valuta', 0) or 0) for line in linii), 2)
    ws2.append(['RO', len([line for line in linii if (line.get('cantitate_ro') or 0) > 0]), int(total_ro), val_ro])
    ws2.append(['HU', len([line for line in linii if (line.get('cantitate_export') or 0) > 0]), int(total_hu), val_hu])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_comenzi_basilur(comanda_id: int):
    """Replică formatul Excel Order Form Basilur PFI."""
    from io import BytesIO
    from db import query, query_one
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from datetime import date

    cmd = query_one("SELECT * FROM comenzi_furnizori WHERE id=?", (comanda_id,))
    if not cmd:
        raise ValueError(f"Comanda {comanda_id} nu există")

    linii = query("""
        SELECT l.*, COALESCE(p.descriere, l.descriere, l.sku) AS denumire
        FROM comenzi_furnizori_linii l
        LEFT JOIN produse p ON p.sku = l.sku
        WHERE l.comanda_id = ?
        ORDER BY l.sku
    """, (comanda_id,))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Order Form'

    ws['A1'] = 'ORDER FORM — BASILUR TEA'
    ws['A2'] = f"Order No: {cmd['nr_comanda']}"
    ws['A3'] = f"Date: {cmd['data_comanda'] or date.today().isoformat()}"
    ws['A4'] = f"ETA: {cmd['data_estimata_livrare'] or ''}"

    col_headers = ['CODE', 'PRODUCT DESCRIPTION', 'UNITS/CTN', 'RO', 'HU', 'TOTAL', 'UNIT PRICE USD', 'TOTAL USD']
    hdr_fill = PatternFill('solid', fgColor='1E3A8A')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(col_headers, 1):
        c = ws.cell(row=14, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font

    total_usd = 0.0
    for i, line in enumerate(linii, 15):
        cant_ro = line.get('cantitate_ro', 0) or 0
        cant_hu = line.get('cantitate_export', 0) or 0
        total   = cant_ro + cant_hu
        pret    = line.get('pret_valuta', 0) or 0
        tot_usd = total * pret
        total_usd += tot_usd
        for col, val in enumerate([
            line.get('cod_furnizor', line['sku']), line['denumire'],
            line.get('units_per_carton', ''), cant_ro, cant_hu, total,
            pret, round(tot_usd, 2),
        ], 1):
            ws.cell(row=i, column=col, value=val)

    footer_row = 15 + len(linii)
    ws.cell(row=footer_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=footer_row, column=8, value=round(total_usd, 2)).font = Font(bold=True)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_expirare(furnizor=None, prag_luni: int = 6):
    """3 sheet-uri: >6 luni, >12 luni, >18 luni."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from queries import expirare_list

    CULORI = {6: 'FFF3CD', 12: 'FFD9B3', 18: 'F8D7DA'}
    headers = ['SKU', 'Cod produs', 'Brand', 'Gamă', 'Cantitate',
               'Data intrare', 'Vechime (zile)', 'Valoare RON', 'Risc']

    wb = openpyxl.Workbook()
    for i, luni in enumerate([6, 12, 18]):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = f'Peste {luni} luni'
        hdr_fill = PatternFill('solid', fgColor='1E3A8A')
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hdr_fill
            c.font = Font(color='FFFFFF', bold=True)
        rows = expirare_list(furnizor=furnizor, prag_luni=luni)
        row_fill = PatternFill('solid', fgColor=CULORI[luni])
        for ri, r in enumerate(rows, 2):
            vals = [r['sku'], r['cod_produs'], r['furnizor'], r['gama'],
                    r['cantitate'], r['data_intrare'], r['vechime_zile'],
                    round(r.get('valoare', 0) or 0, 2),
                    f'+{luni} luni']
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.fill = row_fill

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
