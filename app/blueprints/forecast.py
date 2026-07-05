import json
import logging
import datetime
from flask import Blueprint, render_template, request, jsonify, abort
import queries
import db
from forecast import forecast_logic
from exports.excel_export import send_excel, timestamped_filename

logger = logging.getLogger(__name__)

forecast_bp = Blueprint('forecast', __name__)

MONTHS_RO = ['Ian', 'Feb', 'Mar', 'Apr', 'Mai', 'Iun',
             'Iul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


# ---------------------------------------------------------------------------
# Modul Forecast + Aprovizionare
# ---------------------------------------------------------------------------

@forecast_bp.route('/forecast')
def forecast():
    tab      = request.args.get('tab', 'stoc')
    gama     = request.args.get('gama', '').strip() or None
    urgenta  = request.args.get('urgenta', '').strip() or None
    furnizor = request.args.get('brand', '').strip() or None
    search   = request.args.get('q', '').strip() or None
    sel_status = request.args.get('status', '').strip() or None

    rows       = queries.forecast_stoc_extended(furnizor=furnizor, gama=gama, urgenta=urgenta, search=search)

    summary    = queries.forecast_summary() or {}
    gama_opts  = queries.forecast_gama_list()
    brand_opts = queries.forecast_brands_list()
    lead_times = queries.termene_aprovizionare_list()
    comenzi    = queries.comenzi_list(furnizor=furnizor or None, status=sel_status or None)

    from db import query_one as _qone
    snap = _qone("SELECT MAX(data_snapshot) AS d FROM stoc")
    stoc_snapshot = (snap or {}).get('d') if snap else None

    lt_map = {r['furnizor']: r for r in lead_times}
    piete_active = queries.piete_export_active()

    return render_template(
        'forecast.html',
        tab=tab,
        rows=rows,
        piete_active=piete_active,
        summary=summary,
        gama_opts=gama_opts,
        brand_opts=brand_opts,
        lead_times=lead_times,
        lt_map=lt_map,
        comenzi=comenzi,
        sel_gama=gama or '',
        sel_urgenta=urgenta or '',
        sel_brand=furnizor or '',
        sel_status=sel_status or '',
        sel_search=search or '',
        is_xmas_window=forecast_logic.is_xmas_window(),
        months_json=json.dumps(MONTHS_RO),
        stoc_snapshot=stoc_snapshot,
        today=datetime.date.today().strftime('%d.%m.%Y'),
    )


@forecast_bp.route('/testare')
def testare():
    from feature_flags import SHOW_TESTING
    if not SHOW_TESTING:
        abort(404)
    return render_template('testing_checklist.html')


@forecast_bp.route('/decizii')
def decizii():
    from feature_flags import SHOW_TESTING
    if not SHOW_TESTING:
        abort(404)
    return render_template('decision_torb.html')


@forecast_bp.route('/forecast/setari')
def forecast_setari():
    from forecast import config as fc_config
    tari    = queries.tari_export_list()
    clienti = queries.clienti_export_list()
    termene = queries.termene_aprovizionare_list()
    clienti_by_tara = {}
    for c in clienti:
        clienti_by_tara.setdefault(c['tara_id'], []).append(c)
    return render_template('forecast_setari.html',
        tari=tari, clienti_by_tara=clienti_by_tara, termene=termene,
        params=fc_config.get_params())


@forecast_bp.route('/api/forecast/config', methods=['GET'])
def api_forecast_config_get():
    from forecast import config as fc_config
    return jsonify({'ok': True, 'params': fc_config.get_params()})


@forecast_bp.route('/api/forecast/config', methods=['POST'])
def api_forecast_config_set():
    from forecast import config as fc_config
    d = request.get_json(silent=True) or {}
    try:
        fc_config.set_param(d['cheie'], float(d['valoare']))
        return jsonify({'ok': True})
    except (KeyError, TypeError, ValueError) as e:
        logger.exception("api_forecast_config_set failed")
        return jsonify({'error': str(e)}), 400


@forecast_bp.route('/api/forecast/tari', methods=['POST'])
def api_forecast_tara_save():
    d = request.json or {}
    try:
        queries.tari_export_upsert(
            tara=d['tara'], piata=d['piata'],
            activ=int(d.get('activ', 1)),
            observatii=d.get('observatii'),
            id=d.get('id'),
        )
        return jsonify({'ok': True})
    except (KeyError, TypeError) as e:
        logger.exception("api_forecast_tara_save failed")
        return jsonify({'error': str(e)}), 400


@forecast_bp.route('/api/forecast/tari/<int:id>', methods=['DELETE'])
def api_forecast_tara_delete(id):
    queries.tari_export_delete(id)
    return jsonify({'ok': True})


@forecast_bp.route('/api/forecast/clienti', methods=['POST'])
def api_forecast_client_save():
    d = request.json or {}
    try:
        queries.clienti_export_upsert(
            tara_id=int(d['tara_id']),
            cod_client=d['cod_client'],
            nume_client=d['nume_client'],
            activ=int(d.get('activ', 1)),
            observatii=d.get('observatii'),
            id=d.get('id'),
        )
        return jsonify({'ok': True})
    except (KeyError, TypeError) as e:
        logger.exception("api_forecast_client_save failed")
        return jsonify({'error': str(e)}), 400


@forecast_bp.route('/api/forecast/clienti/<int:id>/toggle', methods=['POST'])
def api_forecast_client_toggle(id):
    queries.clienti_export_toggle(id)
    return jsonify({'ok': True})


@forecast_bp.route('/api/forecast/termene', methods=['POST'])
def api_forecast_termene_save():
    d = request.json or {}
    try:
        queries.termene_aprovizionare_upsert(
            furnizor=d['furnizor'],
            zile_min=int(d.get('zile_min', 30)),
            zile_max=int(d.get('zile_max', 60)),
            moneda=d.get('moneda', 'EUR'),
            tip_produs=d.get('tip_produs', 'Altele'),
            sezon_craciun=int(d.get('sezon_craciun', 0)),
            observatii=d.get('observatii'),
        )
        return jsonify({'ok': True})
    except (KeyError, TypeError) as e:
        logger.exception("api_forecast_termene_save failed")
        return jsonify({'error': str(e)}), 400


# ── Forecast API ──────────────────────────────────────────────────────────────

@forecast_bp.route('/api/forecast/suggest/<path:furnizor>')
def api_forecast_suggest(furnizor):
    try:
        min_velocity = float(request.args.get('min_velocity', 1))
        only_needed = request.args.get('only_needed', '1') == '1'
        result = forecast_logic.build_suggestion(furnizor, min_velocity=min_velocity,
                                                 only_needed=only_needed)
        return jsonify({'ok': True, **result})
    except Exception as exc:
        logger.exception("api_forecast_suggest failed for %s", furnizor)
        return jsonify({'error': str(exc)}), 500


@forecast_bp.route('/api/forecast/sku-clients/<path:sku>')
def api_forecast_sku_clients(sku):
    data = queries.sku_clients_monthly(sku)
    return jsonify({'ok': True, 'sku': sku, 'clients': data})


# ── Comenzi CRUD ──────────────────────────────────────────────────────────────

@forecast_bp.route('/api/comenzi/drafts')
def api_comenzi_drafts():
    """Lista comenzilor draft (pentru dropdown din modulul Stoc)."""
    furnizor = request.args.get('furnizor', '').strip() or None
    rows = queries.comenzi_list(furnizor=furnizor, status='draft')
    return jsonify({'ok': True, 'items': [dict(r) for r in rows]})


@forecast_bp.route('/api/comenzi', methods=['POST'])
def api_comanda_create():
    d = request.get_json(silent=True) or {}
    try:
        cid = queries.comanda_create(d['furnizor'], d.get('nr_comanda'), d.get('observatii'))
        for line in d.get('lines', []):
            queries.comanda_line_upsert(
                cid, line['sku'], int(line.get('cantitate_comandata', 0)),
                int(line.get('cantitate_sugerat', 0)),
                line.get('pret_valuta'), line.get('moneda', 'EUR'), line.get('observatii'),
                cantitate_ro=int(line.get('cantitate_ro', 0)),
                cantitate_export=int(line.get('cantitate_export', 0)),
                cod_furnizor=line.get('cod_furnizor'),
                cantitati_piete=line.get('cantitati_piete'),
            )
        if d.get('status') and d['status'] != 'draft':
            queries.comanda_update(cid, status=d['status'])
        return jsonify({'ok': True, 'id': cid})
    except Exception as exc:
        logger.exception("api_comanda_create failed")
        return jsonify({'error': str(exc)}), 400


# ── Clienți Export (configurabili pentru sugestia de comandă export) ──────
@forecast_bp.route('/api/clienti-export', methods=['GET'])
def api_clienti_export_list():
    rows = db.query("""
        SELECT ce.id, ce.cod_client, ce.nume_client AS client,
               COALESCE(te.piata, 'HU') AS tara,
               ce.activ, ce.observatii
        FROM clienti_export ce
        LEFT JOIN tari_export te ON ce.tara_id = te.id
        ORDER BY ce.activ DESC, ce.nume_client
    """)
    return jsonify({'ok': True, 'items': [dict(r) for r in rows]})


@forecast_bp.route('/api/clienti-export', methods=['POST'])
def api_clienti_export_add():
    d = request.get_json(silent=True) or {}
    try:
        cod = str(d['cod_client']).strip()
        exists = db.query_one("SELECT 1 FROM tranzactii WHERE cod_client=:c LIMIT 1", {'c': cod})
        if not exists:
            return jsonify({'error': f'Codul de client "{cod}" nu apare în nicio tranzacție.'}), 400
        client = (d.get('client') or '').strip()
        if not client:
            r = db.query_one("SELECT client FROM tranzactii WHERE cod_client=:c LIMIT 1", {'c': cod})
            client = r['client'] if r else f'Client {cod}'
        tara_str = (d.get('tara') or 'HU').strip()
        activ = 1 if d.get('activ', True) else 0
        obs = d.get('observatii')
        # Găsim tara_id după piata (HU/RO) sau creăm intrare nouă
        tara_row = db.query_one(
            "SELECT id FROM tari_export WHERE piata=:p LIMIT 1", {'p': tara_str}
        )
        tara_id = tara_row['id'] if tara_row else 1
        conn = db.get_db()
        try:
            conn.execute("""
                INSERT INTO clienti_export (tara_id, cod_client, nume_client, activ, observatii)
                VALUES (?,?,?,?,?)
                ON CONFLICT(cod_client) DO UPDATE SET
                    tara_id=excluded.tara_id, nume_client=excluded.nume_client,
                    activ=excluded.activ, observatii=excluded.observatii
            """, (tara_id, cod, client, activ, obs))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'ok': True})
    except Exception as exc:
        logger.exception("api_clienti_export_add failed")
        return jsonify({'error': str(exc)}), 400


@forecast_bp.route('/api/clienti-export/<cod>', methods=['DELETE'])
def api_clienti_export_delete(cod):
    conn = db.get_db()
    try:
        conn.execute("DELETE FROM clienti_export WHERE cod_client=?", (str(cod),))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


@forecast_bp.route('/api/clienti/search')
def api_clienti_search():
    """Autocompletare clienți după cod sau denumire — pentru tab-ul Export."""
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'ok': True, 'items': []})
    from db import query
    rows = query("""
        SELECT DISTINCT cod_client, client FROM tranzactii
        WHERE (CAST(cod_client AS TEXT) LIKE :q OR client LIKE :q)
          AND cod_client IS NOT NULL
        ORDER BY client LIMIT 15
    """, {'q': f'%{q}%'})
    return jsonify({'ok': True, 'items': [dict(r) for r in rows]})


@forecast_bp.route('/api/comenzi/<int:cid>', methods=['GET'])
def api_comanda_get(cid):
    data = queries.comanda_get(cid)
    if not data:
        return jsonify({'error': 'Comandă negăsită'}), 404
    return jsonify({'ok': True, **data})


@forecast_bp.route('/api/comenzi/<int:cid>', methods=['PUT'])
def api_comanda_update(cid):
    d = request.get_json(silent=True) or {}
    try:
        queries.comanda_update(cid, **d)
        return jsonify({'ok': True})
    except Exception as exc:
        logger.exception("api_comanda_update cid=%s failed", cid)
        return jsonify({'error': str(exc)}), 400


@forecast_bp.route('/api/comenzi/<int:cid>', methods=['DELETE'])
def api_comanda_delete(cid):
    queries.comanda_delete(cid)
    return jsonify({'ok': True})


@forecast_bp.route('/api/comenzi/<int:cid>/lines', methods=['POST'])
def api_comanda_line_add(cid):
    d = request.get_json(silent=True) or {}
    try:
        lid = queries.comanda_line_upsert(
            cid, d['sku'], int(d.get('cantitate_comandata', 0)),
            int(d.get('cantitate_sugerat', 0)),
            d.get('pret_valuta'), d.get('moneda', 'EUR'), d.get('observatii'),
            cantitate_ro=int(d.get('cantitate_ro', 0)),
            cantitate_export=int(d.get('cantitate_export', 0)),
            cod_furnizor=d.get('cod_furnizor'),
            cantitati_piete=d.get('cantitati_piete'),
        )
        return jsonify({'ok': True, 'id': lid})
    except Exception as exc:
        logger.exception("api_comanda_line_add cid=%s failed", cid)
        return jsonify({'error': str(exc)}), 400


@forecast_bp.route('/api/comenzi/<int:cid>/lines/<int:lid>', methods=['PUT'])
def api_comanda_line_update(cid, lid):
    d = request.get_json(silent=True) or {}
    try:
        queries.comanda_line_update(lid, **d)
        return jsonify({'ok': True})
    except Exception as exc:
        logger.exception("api_comanda_line_update lid=%s failed", lid)
        return jsonify({'error': str(exc)}), 400


@forecast_bp.route('/api/comenzi/<int:cid>/lines/<int:lid>', methods=['DELETE'])
def api_comanda_line_delete(cid, lid):
    queries.comanda_line_delete(lid)
    return jsonify({'ok': True})


@forecast_bp.route('/api/comenzi/<int:cid>/status', methods=['POST'])
def api_comanda_status(cid):
    d = request.get_json(silent=True) or {}
    try:
        queries.comanda_update(
            cid,
            status=d.get('status'),
            data_estimata_livrare=d.get('data_estimata_livrare'),
            data_confirmare_furnizor=d.get('data_confirmare_furnizor'),
            observatii=d.get('observatii'),
        )
        for lu in d.get('line_updates', []):
            queries.comanda_line_update(
                lu['id'],
                cantitate_confirmata=lu.get('cantitate_confirmata'),
                cantitate_comandata=lu.get('cantitate_comandata'),
            )
        return jsonify({'ok': True})
    except Exception as exc:
        logger.exception("api_comanda_status cid=%s failed", cid)
        return jsonify({'error': str(exc)}), 400


# ── Termene aprovizionare ─────────────────────────────────────────────────────

@forecast_bp.route('/api/termene-aprovizionare', methods=['POST'])
def api_termene_upsert():
    d = request.get_json(silent=True) or {}
    try:
        queries.termene_partial_update(
            d['furnizor'], int(d['zile_livrare']),
            int(d.get('sezon_craciun', 0)), d.get('observatii'),
        )
        return jsonify({'ok': True})
    except Exception as exc:
        logger.exception("api_termene_upsert failed")
        return jsonify({'error': str(exc)}), 400


# ── Export/Import comenzi ─────────────────────────────────────────────────────

@forecast_bp.route('/export/forecast/comanda/<int:cid>')
def export_comanda(cid):
    """Export comandă în format compatibil cu Basilur Order Form (paste-ready):
    aceleași etichete de coloane ca în PFI (CODE, PRODUCT DESCRIPTION,
    Units per Export, RO, No of Units, Unit Price US$, Total Price US$, etc.)
    plus coloane interne (sku Torb, cantitate sugerat, observații)."""
    data = queries.comanda_get(cid)
    if not data:
        abort(404)
    h = data['header']
    rows = []
    total_units = 0
    total_baxuri = 0
    total_value = 0.0
    total_qty_ro = 0
    total_qty_export = 0
    for line in data['lines']:
        units = line.get('cantitate_comandata') or 0
        baxuri = line.get('cantitate_baxuri') or 0
        upc = line.get('units_per_carton') or 0
        if not baxuri and units and upc:
            baxuri = round(units / upc, 2)
        unit_price = line.get('pret_valuta') or 0
        total_line = line.get('total_valuta')
        if total_line is None and unit_price and units:
            total_line = round(unit_price * units, 2)

        qty_ro = line.get('cantitate_ro') or 0
        qty_export = line.get('cantitate_export') or 0

        total_units  += units or 0
        total_baxuri += baxuri or 0
        total_value  += total_line or 0
        total_qty_ro += qty_ro
        total_qty_export += qty_export

        rows.append({
            'CODE':                 line.get('cod_furnizor') or '',
            'PRODUCT DESCRIPTION':  line.get('descriere') or line.get('sku', ''),
            'Units per Export':     upc or '',
            'RO':                   baxuri or '',
            'Export Ctns.':         baxuri or '',
            'No of Units':          units,
            'Unit Price US$':       unit_price or '',
            'Total Price US$':      total_line or '',
            'Gross Kgs':            line.get('gross_kg') or '',
            'Net Kgs':              line.get('net_kg') or '',
            'CBM':                  line.get('cbm') or '',
            '— Cod intern':         line.get('cod_produs') or '',
            '— SKU intern':         line.get('sku') or '',
            '— Cantitate sugerat':  line.get('cantitate_sugerat') or 0,
            '— Cantitate comandată': line.get('cantitate_comandata') or 0,
            '— Qty piața RO':       qty_ro,
            '— Qty Export HU':      qty_export,
            '— Cantitate confirmată': line.get('cantitate_confirmata') or '',
            '— Observații':         line.get('observatii') or '',
        })

    rows.append({
        'CODE': '',
        'PRODUCT DESCRIPTION': 'TOTAL',
        'Units per Export': '',
        'RO': total_baxuri,
        'Export Ctns.': total_baxuri,
        'No of Units': total_units,
        'Unit Price US$': '',
        'Total Price US$': round(total_value, 2),
        'Gross Kgs': '', 'Net Kgs': '', 'CBM': '',
        '— Cod intern': '', '— SKU intern': '', '— Cantitate sugerat': '', '— Cantitate comandată': '',
        '— Qty piața RO': total_qty_ro,
        '— Qty Export HU': total_qty_export,
        '— Cantitate confirmată': '', '— Observații': '',
    })

    fname = f"comanda_{h['furnizor']}_{h['data_comanda']}".replace(' ', '_')
    return send_excel({'Order form': rows}, timestamped_filename(fname))


@forecast_bp.route('/import/forecast/comanda/<int:cid>', methods=['POST'])
def import_comanda_lines(cid):
    if 'file' not in request.files:
        return jsonify({'error': 'Fișier lipsă'}), 400
    f = request.files['file']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value or '').upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        sku_col = next((i for i, h in enumerate(headers) if 'SKU' in h), None)
        qty_col = next((i for i, h in enumerate(headers) if 'COMAND' in h), None)
        if sku_col is None or qty_col is None:
            return jsonify({'error': 'Nu am găsit coloanele SKU și Cantitate Comandat'}), 400
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = row[sku_col] if sku_col < len(row) else None
            qty = row[qty_col] if qty_col < len(row) else None
            if sku and qty:
                try:
                    queries.comanda_line_upsert(cid, str(sku).strip(), int(float(qty)))
                    count += 1
                except Exception:
                    logger.warning("comanda_line_upsert skipped sku=%s qty=%s", sku, qty, exc_info=True)
        return jsonify({'ok': True, 'imported': count})
    except Exception as exc:
        logger.exception("comanda Excel import failed for cid=%s", cid)
        return jsonify({'error': str(exc)}), 400


# ── Forecast agent chat ───────────────────────────────────────────────────────

@forecast_bp.route('/api/forecast/chat', methods=['POST'])
def api_forecast_chat():
    d = request.get_json(silent=True) or {}
    question = (d.get('question') or '').strip()
    furnizor = (d.get('furnizor') or '').strip() or None
    if not question:
        return jsonify({'error': 'Întrebarea lipsă'}), 400
    try:
        from forecast.forecast_agent import forecast_ask
        result = forecast_ask(question, furnizor)
        return jsonify(result)
    except Exception as exc:
        logger.exception("api_forecast_chat failed")
        return jsonify({'error': str(exc)}), 500


