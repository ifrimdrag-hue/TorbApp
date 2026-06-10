"""Excel export of forecast run results.

4 sheets:
    1. Reorder       — SKU-level reorder recommendations, sorted by urgency.
    2. Forecast_Luna — brand × canal × month cashflow view.
    3. Alerts        — critical/high urgency items and top risks.
    4. Metodologie   — parameters, brand config, data freshness, how to read.
"""

import io
import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import DB_PATH


URGENCY_COLORS = {
    "critical": "FFC7CE",  # light red
    "high":     "FFEB9C",  # light yellow
    "normal":   "C6EFCE",  # light green
    "unknown":  "D9D9D9",  # light gray
}
URGENCY_ORDER = {"critical": 0, "high": 1, "unknown": 2, "normal": 3}


def _latest_run_id(conn):
    row = conn.execute(
        "SELECT run_id FROM forecast_runs WHERE status='done' "
        "ORDER BY run_id DESC LIMIT 1"
    ).fetchone()
    if not row:
        raise ValueError("No completed forecast run found. Run forecast.run first.")
    return row[0]


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F5D75", end_color="4F5D75",
                                 fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width=60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0
                       for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10),
                                                      max_width)


def _write_reorder_sheet(wb, conn, run_id):
    df = pd.read_sql_query("""
        SELECT furnizor, cod_produs, sku, stock_on_hand,
               demand_over_lead_time, safety_stock, reorder_point,
               suggested_qty, order_by_date, urgency, rationale
        FROM reorder_suggestions WHERE run_id = ?
    """, conn, params=(run_id,))

    df["urg_rank"] = df["urgency"].map(URGENCY_ORDER).fillna(99)
    df = df.sort_values(["urg_rank", "demand_over_lead_time"],
                        ascending=[True, False]).drop(columns=["urg_rank"])

    ws = wb.create_sheet("Reorder")
    ws.append([
        "Furnizor", "Cod produs", "SKU", "Stoc curent",
        "Cerere orizont (buc)", "Safety stock", "Reorder point",
        "Cantitate sugerată", "Comandă până la", "Urgență", "Note",
    ])
    _style_header(ws)

    for r in df.itertuples(index=False):
        ws.append([
            r.furnizor, r.cod_produs, r.sku,
            r.stock_on_hand,
            r.demand_over_lead_time, r.safety_stock, r.reorder_point,
            r.suggested_qty, r.order_by_date, r.urgency, r.rationale,
        ])
        color = URGENCY_COLORS.get(r.urgency, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color,
                           fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_monthly_sheet(wb, conn, run_id):
    df = pd.read_sql_query("""
        SELECT furnizor, canal, week_start, yhat
        FROM forecasts WHERE run_id = ?
    """, conn, params=(run_id,))
    if df.empty:
        return
    df["week_start"] = pd.to_datetime(df["week_start"])
    df["luna"] = df["week_start"].dt.to_period("M").astype(str)

    pivot = (df.groupby(["furnizor", "canal", "luna"], as_index=False)["yhat"]
               .sum()
               .pivot(index=["furnizor", "canal"], columns="luna",
                       values="yhat")
               .reset_index()
               .fillna(0))

    ws = wb.create_sheet("Forecast_Luna")
    ws.append(list(pivot.columns))
    _style_header(ws)
    for r in pivot.itertuples(index=False):
        row = [r[0], r[1]] + [round(float(v), 0) for v in r[2:]]
        ws.append(row)
    ws.freeze_panes = "C2"
    _autosize(ws, max_width=20)


def _write_alerts_sheet(wb, conn, run_id):
    critical = pd.read_sql_query("""
        SELECT furnizor, cod_produs, sku, stock_on_hand,
               demand_over_lead_time, safety_stock, order_by_date, urgency
        FROM reorder_suggestions
        WHERE run_id = ? AND urgency IN ('critical', 'high')
        ORDER BY urgency, demand_over_lead_time DESC
    """, conn, params=(run_id,))

    ws = wb.create_sheet("Alerts")
    ws.append(["Tip", "Furnizor", "Cod produs", "SKU",
                "Stoc", "Cerere orizont", "Safety stock",
                "Comandă până la"])
    _style_header(ws)
    for r in critical.itertuples(index=False):
        ws.append([
            "REORDER " + r.urgency.upper(),
            r.furnizor, r.cod_produs, r.sku,
            r.stock_on_hand, r.demand_over_lead_time, r.safety_stock,
            r.order_by_date,
        ])
        color = URGENCY_COLORS.get(r.urgency, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color,
                           fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # SKU with no sales in last 12 weeks (potentially dying)
    dying = pd.read_sql_query("""
        SELECT DISTINCT t.furnizor, t.cod_produs, t.sku,
               MAX(t.data_dl) AS ultima_vanzare
        FROM tranzactii t
        WHERE t.data_dl < DATE((SELECT MAX(data_dl) FROM tranzactii),
                               '-84 days')
          AND t.cod_produs NOT IN (
            SELECT cod_produs FROM tranzactii
            WHERE data_dl >= DATE((SELECT MAX(data_dl) FROM tranzactii),
                                   '-84 days')
          )
        GROUP BY t.cod_produs
        ORDER BY ultima_vanzare DESC
        LIMIT 50
    """, conn)
    if not dying.empty:
        ws.append([])  # blank row
        ws.append(["SKU fără vânzări în ultimele 12 săptămâni (potențial delistare):"])
        for r in dying.itertuples(index=False):
            ws.append(["DYING", r.furnizor, r.cod_produs, r.sku, "", "", "",
                        r.ultima_vanzare])

    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_methodology_sheet(wb, conn, run_id):
    meta = conn.execute(
        "SELECT * FROM forecast_runs WHERE run_id = ?", (run_id,)
    ).fetchone()
    cols = [d[0] for d in conn.execute(
        "SELECT * FROM forecast_runs LIMIT 0"
    ).description]
    meta = dict(zip(cols, meta)) if meta else {}

    brand_cfg = pd.read_sql_query("SELECT * FROM brands_config", conn)
    stock_freshness = conn.execute(
        "SELECT MAX(snapshot_date) FROM stock_snapshot"
    ).fetchone()[0] or "— încă nu s-a încărcat stocul —"

    ws = wb.create_sheet("Metodologie")
    ws.append(["Parametru", "Valoare"])
    _style_header(ws)
    ws.append(["Run ID", meta.get("run_id")])
    ws.append(["Status", meta.get("status")])
    ws.append(["Pornit la", meta.get("started_at")])
    ws.append(["Finalizat la", meta.get("finished_at")])
    ws.append(["Orizont (săptămâni)", meta.get("horizon_weeks")])
    ws.append(["Branduri", meta.get("brands_included")])
    ws.append(["Input hash", meta.get("input_hash")])
    ws.append(["Stoc (ultima actualizare)", stock_freshness])
    ws.append([])
    ws.append(["Brand config:"])
    ws.append(list(brand_cfg.columns))
    for r in brand_cfg.itertuples(index=False):
        ws.append(list(r))
    ws.append([])
    ws.append(["Metodologie:"])
    ws.append(["1. Forecast la nivel brand × canal × săptămână (AutoETS)."])
    ws.append(["2. Q4 (Oct-Dec) aplicat ca overlay multiplicativ per brand."])
    ws.append(["3. Vara (Jun-Aug) dampener pentru Toras/Delaviuda (fără depozit frig)."])
    ws.append(["4. Alocare SKU pe share rulant 12 săpt, reconciliere middle-out."])
    ws.append(["5. Safety stock = z × σ_săptămânal × sqrt(lead+review)."])
    ws.append(["6. Reorder = cerere pe lead_time + safety_stock."])
    ws.append(["7. Cantitate sugerată = target - (stoc + în_comandă), rotunjit la MOQ."])
    _autosize(ws, max_width=50)


def build_workbook(run_id=None, db_path=DB_PATH):
    with sqlite3.connect(db_path) as conn:
        if run_id is None:
            run_id = _latest_run_id(conn)

        wb = Workbook()
        # Remove default sheet — we create our own.
        default = wb.active
        wb.remove(default)

        _write_reorder_sheet(wb, conn, run_id)
        _write_monthly_sheet(wb, conn, run_id)
        _write_alerts_sheet(wb, conn, run_id)
        _write_methodology_sheet(wb, conn, run_id)

    return wb


def export_to_file(path, run_id=None, db_path=DB_PATH):
    wb = build_workbook(run_id=run_id, db_path=db_path)
    wb.save(path)
    return path


def export_to_bytes(run_id=None, db_path=DB_PATH):
    wb = build_workbook(run_id=run_id, db_path=db_path)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run", type=int, help="run_id (default: latest done)")
    parser.add_argument("--out", default=None,
                        help="output path (default: forecast_<run_id>.xlsx)")
    args = parser.parse_args()

    if not args.out:
        args.out = f"forecast_{args.run or 'latest'}.xlsx"
    path = export_to_file(args.out, run_id=args.run)
    print(f"wrote {path}")
