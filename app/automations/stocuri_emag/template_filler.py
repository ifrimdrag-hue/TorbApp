"""Genereaza fisierul eMAG completat cu stocurile corecte.

Flow:
  1. Incarcam template-ul (export-ul eMAG) cu openpyxl, pastrand formatarea
  2. Localizam coloanele 'ean' si 'stock' din header-ul de pe randul 3 al sheet-ului 'Oferte'
  3. Pentru fiecare rand de date (incepand cu randul 5):
     - Daca EAN-ul exista in raportul intern → scriem cantitatea
     - Daca nu → scriem 0 (produs nu mai e in stoc)
  4. Salvam in BytesIO si returnam bytes-urile
"""

from io import BytesIO
from typing import NamedTuple
from openpyxl import load_workbook


SHEET_NAME = "Oferte"
HEADER_ROW = 3
DATA_START_ROW = 5
EAN_HEADER = "ean"
STOCK_HEADER = "stock"


class TemplateFillResult(NamedTuple):
    file_bytes: bytes
    matched: int                       # randuri eMAG mapate cu stoc din raport
    set_to_zero: int                   # randuri eMAG fara match → puse 0
    template_rows_no_ean: int          # randuri eMAG fara EAN (set la 0)
    matched_eans: set[str]             # EAN-urile efectiv folosite din raport


def _normalize_ean(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN check
            return None
        if value.is_integer():
            return str(int(value))
        return None
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def fill_emag_template(
    template_bytes: bytes,
    stocks_by_ean: dict[str, int],
) -> TemplateFillResult:
    wb = load_workbook(BytesIO(template_bytes))
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Template-ul eMAG nu are sheet-ul '{SHEET_NAME}'. "
            f"Sheet-uri gasite: {', '.join(wb.sheetnames)}"
        )
    ws = wb[SHEET_NAME]

    # Localizeaza coloanele in header
    headers: dict[str, int] = {}
    for cell in ws[HEADER_ROW]:
        if cell.value is not None:
            key = str(cell.value).strip().lower()
            if key:
                headers[key] = cell.column

    if EAN_HEADER not in headers:
        raise ValueError(
            f"Coloana '{EAN_HEADER}' nu a fost gasita pe randul {HEADER_ROW} din sheet-ul '{SHEET_NAME}'. "
            f"Coloane gasite: {', '.join(headers.keys())}"
        )
    if STOCK_HEADER not in headers:
        raise ValueError(
            f"Coloana '{STOCK_HEADER}' nu a fost gasita pe randul {HEADER_ROW} din sheet-ul '{SHEET_NAME}'."
        )

    ean_col = headers[EAN_HEADER]
    stock_col = headers[STOCK_HEADER]

    matched = 0
    set_to_zero = 0
    template_rows_no_ean = 0
    matched_eans: set[str] = set()

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            continue

        ean_cell = row[ean_col - 1]
        stock_cell = row[stock_col - 1]
        ean = _normalize_ean(ean_cell.value)

        if ean is None:
            template_rows_no_ean += 1
            stock_cell.value = 0
            set_to_zero += 1
            continue

        qty = stocks_by_ean.get(ean)
        if qty is None:
            stock_cell.value = 0
            set_to_zero += 1
        else:
            stock_cell.value = int(qty)
            matched += 1
            matched_eans.add(ean)

    out = BytesIO()
    wb.save(out)
    return TemplateFillResult(
        file_bytes=out.getvalue(),
        matched=matched,
        set_to_zero=set_to_zero,
        template_rows_no_ean=template_rows_no_ean,
        matched_eans=matched_eans,
    )
