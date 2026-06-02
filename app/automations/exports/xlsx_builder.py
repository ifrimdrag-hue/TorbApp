"""Excel operational pentru tracking campanii, buget, task-uri si KPI-uri."""

from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference, PieChart

from . import theme as T


def _hdr_font():
    return Font(name=T.FONT_BODY, bold=True, color=T.WHITE, size=11)


def _navy_fill():
    return PatternFill(start_color=T.NAVY, end_color=T.NAVY, fill_type="solid")


def _alt_fill():
    return PatternFill(start_color=T.LIGHT_GRAY, end_color=T.LIGHT_GRAY, fill_type="solid")


def _border():
    s = Side(style="thin", color=T.MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_header_row(ws, row: int, col_count: int):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _hdr_font()
        cell.fill = _navy_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[row].height = 22


def _style_data_rows(ws, start_row: int, end_row: int, col_count: int):
    body_font = Font(name=T.FONT_BODY, size=10, color=T.DARK_GRAY)
    for r in range(start_row, end_row + 1):
        fill = _alt_fill() if (r - start_row) % 2 == 1 else None
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            if fill:
                cell.fill = fill
            cell.border = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 18


def build_xlsx(campaigns: list[dict]) -> bytes:
    wb = Workbook()

    # ──────────── SHEET 1: SUMAR + BUGET ────────────
    ws = wb.active
    ws.title = "Sumar & Buget"

    # Titlu mare
    ws["A1"] = "Plan Campanii Mai 2026 — Sumar Buget"
    ws["A1"].font = Font(name=T.FONT_TITLE, bold=True, size=18, color=T.NAVY)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = f"Generat: {date.today().strftime('%d %B %Y')}"
    ws["A2"].font = Font(name=T.FONT_BODY, italic=True, size=9, color=T.MID_GRAY)
    ws.merge_cells("A2:F2")

    # Header tabel
    headers = ["Campanie", "Tip", "Status", "Perioada", "Buget Alocat (RON)", "Buget Cheltuit (RON)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))

    # Date campanii
    type_labels = {"promo": "Promo", "gifting": "Gifting", "lansare": "Lansare", "sezonier": "Sezonier", "giveaway": "Giveaway"}
    status_labels = {"draft": "Draft", "planned": "Planificata", "active": "Activa", "completed": "Finalizata", "cancelled": "Anulata"}
    row = 5
    for c in campaigns:
        ws.cell(row=row, column=1, value=c.get("name", ""))
        ws.cell(row=row, column=2, value=type_labels.get(c.get("type"), c.get("type", "")))
        ws.cell(row=row, column=3, value=status_labels.get(c.get("status"), c.get("status", "")))
        ws.cell(row=row, column=4, value=f"{c.get('date_start','')} → {c.get('date_end','')}")
        ws.cell(row=row, column=5, value=c.get("budget_alloc") or 0)
        ws.cell(row=row, column=6, value=c.get("budget_spent") or 0)
        row += 1

    # Total
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=5, value=f"=SUM(E5:E{row - 1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F5:F{row - 1})")
    for c in range(1, 7):
        cell = ws.cell(row=total_row, column=c)
        cell.font = Font(name=T.FONT_BODY, bold=True, size=11, color=T.NAVY)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        cell.border = _border()

    _style_data_rows(ws, 5, row - 1, 6)
    _set_col_widths(ws, [40, 14, 14, 26, 18, 18])

    # Format numeric pe coloanele de buget
    for r in range(5, total_row + 1):
        ws.cell(row=r, column=5).number_format = "#,##0 \"RON\""
        ws.cell(row=r, column=6).number_format = "#,##0 \"RON\""

    # Pie chart distributie buget
    if len(campaigns) >= 1:
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=5, max_row=5 + len(campaigns) - 1)
        data = Reference(ws, min_col=5, min_row=4, max_row=5 + len(campaigns) - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distributie buget pe campanii"
        pie.height = 10
        pie.width = 15
        ws.add_chart(pie, f"A{total_row + 3}")

    # ──────────── SHEET 2: CAMPANII DETALIATE ────────────
    ws2 = wb.create_sheet("Detalii Campanii")
    ws2["A1"] = "Detalii complete pe fiecare campanie"
    ws2["A1"].font = Font(name=T.FONT_TITLE, bold=True, size=16, color=T.NAVY)
    ws2.merge_cells("A1:E1")
    ws2.row_dimensions[1].height = 28

    headers2 = ["Campanie", "Mecanica", "Canale", "Produse incluse", "Note"]
    for i, h in enumerate(headers2, start=1):
        ws2.cell(row=3, column=i, value=h)
    _style_header_row(ws2, 3, len(headers2))

    channel_labels = {"shopify": "Shopify", "emag": "eMAG", "instagram": "Instagram", "facebook": "Facebook"}
    row = 4
    for c in campaigns:
        ws2.cell(row=row, column=1, value=c.get("name", ""))
        ws2.cell(row=row, column=2, value=c.get("mechanic", ""))
        channels = ", ".join(channel_labels.get(ch, ch) for ch in c.get("channels", []))
        ws2.cell(row=row, column=3, value=channels)
        prods = c.get("products", []) or []
        prod_summary = f"{len(prods)} produse: " + ", ".join(p.get("name", p.get("sku", "?"))[:30] for p in prods[:4])
        if len(prods) > 4:
            prod_summary += f" + {len(prods) - 4} alte"
        ws2.cell(row=row, column=4, value=prod_summary)
        ws2.cell(row=row, column=5, value=c.get("notes", ""))
        ws2.row_dimensions[row].height = 60
        row += 1

    _style_data_rows(ws2, 4, row - 1, len(headers2))
    _set_col_widths(ws2, [30, 50, 25, 60, 50])

    # ──────────── SHEET 3: TASK-URI (TRACKER) ────────────
    ws3 = wb.create_sheet("Task-uri")
    ws3["A1"] = "Tracker livrabile (livrabile per campanie)"
    ws3["A1"].font = Font(name=T.FONT_TITLE, bold=True, size=16, color=T.NAVY)
    ws3.merge_cells("A1:G1")
    ws3.row_dimensions[1].height = 28

    headers3 = ["Status", "Campanie", "Task", "Prioritate", "Asignat", "Tip", "Deadline"]
    for i, h in enumerate(headers3, start=1):
        ws3.cell(row=3, column=i, value=h)
    _style_header_row(ws3, 3, len(headers3))

    status_icons = {"todo": "○ Todo", "in_progress": "◐ In lucru", "blocked": "⊘ Blocat", "done": "✓ Done"}
    prio_labels = {"low": "Scazuta", "medium": "Medie", "high": "Inalta", "urgent": "Urgent"}
    assignee_type_labels = {"internal": "Intern", "external": "Extern"}

    row = 4
    for c in campaigns:
        for t in c.get("tasks", []) or []:
            ws3.cell(row=row, column=1, value=status_icons.get(t.get("status"), t.get("status", "")))
            ws3.cell(row=row, column=2, value=c.get("name", ""))
            ws3.cell(row=row, column=3, value=t.get("title", ""))
            ws3.cell(row=row, column=4, value=prio_labels.get(t.get("priority"), t.get("priority", "")))
            ws3.cell(row=row, column=5, value=t.get("assignee", ""))
            ws3.cell(row=row, column=6, value=assignee_type_labels.get(t.get("assignee_type"), ""))
            ws3.cell(row=row, column=7, value=t.get("deadline", ""))
            row += 1

    if row > 4:
        _style_data_rows(ws3, 4, row - 1, len(headers3))
    _set_col_widths(ws3, [14, 30, 50, 12, 22, 10, 12])

    # ──────────── SHEET 4: KPI TRACKER (de actualizat saptamanal) ────────────
    ws4 = wb.create_sheet("KPI Tracker")
    ws4["A1"] = "KPI Tracker — actualizat saptamanal"
    ws4["A1"].font = Font(name=T.FONT_TITLE, bold=True, size=16, color=T.NAVY)
    ws4.merge_cells("A1:F1")
    ws4.row_dimensions[1].height = 28

    ws4["A2"] = "Completeaza saptamanal coloanele 'Saptamana N' pentru a urmari progresul"
    ws4["A2"].font = Font(name=T.FONT_BODY, italic=True, size=9, color=T.MID_GRAY)
    ws4.merge_cells("A2:F2")

    headers4 = ["KPI", "Tinta luna", "Sapt 1 (1-7 mai)", "Sapt 2 (8-14)", "Sapt 3 (15-21)", "Sapt 4 (22-31)"]
    for i, h in enumerate(headers4, start=1):
        ws4.cell(row=4, column=i, value=h)
    _style_header_row(ws4, 4, len(headers4))

    kpi_rows = [
        ["Comenzi B2C totale", "350-450", "", "", "", ""],
        ["Revenue B2C atribuibil (RON)", "25.000-40.000", "", "", "", ""],
        ["AOV mediu (RON)", "+15-25% vs aprilie", "", "", "", ""],
        ["Contribution margin %", ">28%", "", "", "", ""],
        ["Bundle-uri Mama Mea vandute", "60-100", "", "", "", ""],
        ["Comenzi Sezonul Schimba", "200+", "", "", "", ""],
        ["Bundle-uri Verile Usor", "80-150", "", "", "", ""],
        ["Engagement rate IG", ">3%", "", "", "", ""],
        ["Email subscribers nou", "+200-400", "", "", "", ""],
        ["Lichidare stoc Torras", "60-70%", "", "", "", ""],
    ]
    for r_idx, row_data in enumerate(kpi_rows, start=5):
        for c_idx, val in enumerate(row_data, start=1):
            ws4.cell(row=r_idx, column=c_idx, value=val)

    _style_data_rows(ws4, 5, 4 + len(kpi_rows), len(headers4))
    _set_col_widths(ws4, [38, 22, 18, 18, 18, 18])

    # ──────────── SHEET 5: RISCURI ────────────
    ws5 = wb.create_sheet("Riscuri")
    ws5["A1"] = "Riscuri & Mitigari"
    ws5["A1"].font = Font(name=T.FONT_TITLE, bold=True, size=16, color=T.NAVY)
    ws5.merge_cells("A1:D1")
    ws5.row_dimensions[1].height = 28

    headers5 = ["Risc", "Probabilitate", "Impact", "Mitigare"]
    for i, h in enumerate(headers5, start=1):
        ws5.cell(row=3, column=i, value=h)
    _style_header_row(ws5, 3, len(headers5))

    risks = [
        ["Stoc insuficient pentru Tea Book de Ziua Mamei", "Medie", "Mare",
         "Verific stoc inainte de lansare prin Hub. Limitez bundle la N buc disponibile."],
        ["Canibalizare intre campania B si C", "Mica", "Mediu",
         "Audiente diferite — B = lifestyle general, C = health-conscious. Mesaje distincte."],
        ["Vreme rece prelungita → 'vara' nu prinde", "Mica-Medie", "Mediu",
         "Mesaj B/C poate fi pivotat la 'indoor coziness'."],
        ["Foto-sesiunea nu se intampla la timp", "Medie", "Mare",
         "Programare imediat dupa aprobarea sedintei. Buffer 2 zile inainte de lansare."],
        ["App Shopify pentru 2+1 nu functioneaza", "Identificat (BLOCAT)", "Mediu",
         "Solutie B: discount cod manual la checkout. Backup deja in plan."],
    ]
    for r_idx, row_data in enumerate(risks, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            ws5.cell(row=r_idx, column=c_idx, value=val)
        ws5.row_dimensions[r_idx].height = 35

    _style_data_rows(ws5, 4, 3 + len(risks), len(headers5))
    _set_col_widths(ws5, [40, 18, 14, 50])

    # Salvare in BytesIO
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
