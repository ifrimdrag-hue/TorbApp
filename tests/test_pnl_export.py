import openpyxl
import db
from exports.excel_export import build_pnl_xlsx


def _seed():
    conn = db.get_db()
    conn.execute("DELETE FROM pnl_balante_raw")
    conn.executemany(
        """INSERT INTO pnl_balante_raw
           (source_file,entitate,an,luna,cont,dencont,sid,sic,sfd,sfc,rulld,rullc,rulcd,rulcc)
           VALUES('f',?,?,?,'707','',0,0,0,0,0,0,?,0)""",
        [('torb', 2025, 1, 1000.0), ('tobra', 2025, 1, 500.0)])
    conn.commit()
    conn.close()


def test_build_pnl_xlsx_sheets():
    _seed()
    buf = build_pnl_xlsx(2025)
    wb = openpyxl.load_workbook(buf)
    assert wb.sheetnames == ['P&L Torb', 'P&L Tobra', 'P&L Grup', 'KPI Summary']
    ws = wb['P&L Torb']
    assert ws['A1'].value == 'Linie P&L'
