"""Pricing F3: client xls files (listing templates + photo offer)."""
from io import BytesIO
import sqlite3

import openpyxl
import pytest


@pytest.fixture()
def propunere_f3(client, db_path):
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        INSERT OR IGNORE INTO produse (sku, descriere, furnizor, categorie,
            gama, gramaj, ean, tva_pct, buc_cutie, activ)
        VALUES ('F3-1', 'Ceai F3', 'Basilur', 'CEAI', 'Basilur',
                100, '590000000001', 0.09, 24, 1);
        INSERT OR IGNORE INTO costuri_landing (an, sku, moneda,
            pret_achizitie_valuta, curs_ron, pret_achizitie_ron,
            transport_pct, taxa_vamala_pct, alte_costuri_ron, landing_cost_ron)
        VALUES (2026, 'F3-1', 'USD', 10, 4.6, 46, 0, 0, 0, 46);
        INSERT OR IGNORE INTO preturi_vanzare (an, sku, cod_client,
            pret_vanzare_ron, activ) VALUES (2026, 'F3-1', 'CL-F3', 60, 1);
        INSERT INTO tranzactii (an, cod_client, client, furnizor, sku)
        SELECT 2026, 'CL-F3', 'KAUFLAND TEST SCS', 'Basilur', 'F3-1'
        WHERE NOT EXISTS (SELECT 1 FROM tranzactii WHERE cod_client='CL-F3');
        INSERT OR IGNORE INTO coduri_client_articol (sku, cod_client, cod_intern)
        VALUES ('F3-1', 'CL-F3', '133482');
        INSERT OR REPLACE INTO clienti_pricing (cod_client, nume_client,
            template_listare) VALUES ('CL-F3', 'KAUFLAND TEST SCS',
            'kaufland_modificare');
        INSERT OR IGNORE INTO produse_logistica (sku, buc_bax, bax_palet)
        VALUES ('F3-1', 24, 96);
    """)
    conn.commit()
    conn.close()
    rv = client.post('/api/preturi/propuneri', json={
        'an': 2026, 'cod_client': 'CL-F3', 'titlu': 'F3 export',
        'linii': [{'sku': 'F3-1', 'pret_propus': 69}]})
    pid = rv.get_json()['id']
    yield pid
    client.delete(f'/api/preturi/propuneri/{pid}')


def _sheet(resp):
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.mimetype
    return openpyxl.load_workbook(BytesIO(resp.data)).active


def test_listare_kaufland_from_client_setting(client, propunere_f3):
    ws = _sheet(client.get(
        f'/preturi/propuneri/{propunere_f3}/listare.xlsx?valabil=15.07.2026'))
    assert ws.cell(4, 1).value == 'Cod articol'
    assert ws.cell(4, 2).value == 'Cod Kaufland'
    row = [ws.cell(5, c).value for c in range(1, 9)]
    assert row == ['F3-1', '133482', 'Ceai F3', 60, 60, 69, 69, '15.07.2026']


def test_listare_template_override(client, propunere_f3):
    ws = _sheet(client.get(
        f'/preturi/propuneri/{propunere_f3}/listare.xlsx?template=fildas_lista'))
    assert ws.cell(3, 1).value == 'Cod art fz'
    assert [ws.cell(5, c).value for c in range(1, 6)] == \
        ['F3-1', 'Ceai F3', '100g', 60, 69]


def test_listare_selgros_bax_math(client, propunere_f3):
    ws = _sheet(client.get(
        f'/preturi/propuneri/{propunere_f3}/listare.xlsx?template=selgros_lista'))
    row = [ws.cell(7, c).value for c in range(1, 13)]
    # UC=24 -> case price 69*24, unit price 69, EAN and pallet count present
    assert row[3] == 24 and row[5] == pytest.approx(1656.0)
    assert row[6] == 69 and row[10] == '590000000001' and row[11] == 96


def test_listare_generic_fallback(client, propunere_f3):
    ws = _sheet(client.get(
        f'/preturi/propuneri/{propunere_f3}/listare.xlsx?template=inexistent'))
    assert 'TORB LOGISTIC' in ws.cell(1, 1).value


def test_oferta_with_prices_and_tva(client, propunere_f3):
    ws = _sheet(client.get(
        f'/preturi/propuneri/{propunere_f3}/oferta.xlsx?valabil=15.07.2026'))
    assert ws.cell(4, 1).value == 'Poza'
    row = [ws.cell(5, c).value for c in range(1, 10)]
    assert row[1] == 'F3-1' and row[6] == 69
    assert row[7] == pytest.approx(9.0)          # TVA %
    assert row[8] == pytest.approx(75.21)        # 69 * 1.09
