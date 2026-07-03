"""Regression tests for app/queries/forecast.py — forecast page fixes."""
import sqlite3
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'app'))


def _conn(db_path):
    return sqlite3.connect(db_path)


def _next_snapshot(conn):
    """Return a data_snapshot date guaranteed newer than any existing row.

    forecast_stoc_extended() and friends filter on a GLOBAL
    MAX(data_snapshot) FROM stoc, not scoped per furnizor/sku. Tests
    sharing the session-scoped db_path must each pick a date newer than
    every prior insert or their rows silently drop out of the "latest
    snapshot" filter — computing it here decouples tests from execution
    order instead of relying on hardcoded increasing literals.
    """
    current = conn.execute("SELECT MAX(data_snapshot) FROM stoc").fetchone()[0]
    if current is None:
        return '2026-07-01'
    return conn.execute("SELECT date(?, '+1 day')", (current,)).fetchone()[0]


def test_forecast_stoc_extended_includes_price(db_path, client):
    conn = _conn(db_path)
    snap = _next_snapshot(conn)
    conn.execute("""
        INSERT INTO stoc (data_snapshot, cod_produs, cod_mare, sku, furnizor, gama,
                           cantitate, pret_achizitie, data_intrare)
        VALUES (?, 'A5-001', 'A5-001', 'SKU-A5-001', 'TestBrandA5', 'Ceai',
                100, 10.0, '2026-06-01')
    """, (snap,))
    conn.execute("""
        INSERT INTO costuri_landing (an, sku, moneda, pret_achizitie_valuta)
        VALUES (2026, 'SKU-A5-001', 'EUR', 3.5)
    """)
    conn.commit()
    conn.close()

    import queries
    rows = queries.forecast_stoc_extended(furnizor='TestBrandA5')
    assert len(rows) == 1
    assert rows[0]['pret_valuta'] == 3.5
    assert rows[0]['moneda_valuta'] == 'EUR'


def test_forecast_summary_counts_skus_not_lots(db_path, client):
    conn = _conn(db_path)
    snap = _next_snapshot(conn)
    # Same SKU, two lots (multi-lot SKU) — must count once, not twice
    conn.execute("""
        INSERT INTO stoc (data_snapshot, cod_produs, cod_mare, sku, furnizor, gama,
                           cantitate, pret_achizitie, data_intrare)
        VALUES (?, 'B1-001', 'B1-001', 'SKU-B1-MULTILOT', 'TestBrandB1', 'Ceai',
                10, 5.0, '2026-05-01')
    """, (snap,))
    conn.execute("""
        INSERT INTO stoc (data_snapshot, cod_produs, cod_mare, sku, furnizor, gama,
                           cantitate, pret_achizitie, data_intrare)
        VALUES (?, 'B1-001', 'B1-001', 'SKU-B1-MULTILOT', 'TestBrandB1', 'Ceai',
                10, 5.0, '2026-06-01')
    """, (snap,))
    conn.commit()
    conn.close()

    import queries
    summary = queries.forecast_summary()
    total_counted = (summary['critic'] or 0) + (summary['atentie'] or 0) + (summary['ok'] or 0)
    assert total_counted == summary['nr_sku'], (
        f"critic+atentie+ok ({total_counted}) must equal nr_sku ({summary['nr_sku']}) "
        "— a multi-lot SKU must count once"
    )


def test_zile_stoc_excludes_transit(db_path, client):
    conn = _conn(db_path)
    snap = _next_snapshot(conn)
    conn.execute("""
        INSERT INTO stoc (data_snapshot, cod_produs, cod_mare, sku, furnizor, gama,
                           cantitate, pret_achizitie, data_intrare)
        VALUES (?, 'B2-001', 'B2-001', 'SKU-B2-001', 'TestBrandB2', 'Ceai',
                30, 10.0, '2026-06-01')
    """, (snap,))
    # 3 years of sales so the 3-year monthly average kicks in and overwrites zile_stoc
    for luna in range(1, 13):
        conn.execute("""
            INSERT INTO tranzactii (luna, an, data_dl, sku, furnizor, cantitate,
                                     cod_produs, client, cod_client, agent,
                                     pret_vanzare, tva_pct, pret_cumparare,
                                     val_bruta, val_neta, val_achizitie, marja_bruta, discount_pct)
            VALUES (:luna, 2025, '2025-' || printf('%02d', :luna) || '-10',
                    'SKU-B2-001', 'TestBrandB2', 30,
                    'B2-001', 'Client Test', 'C001', 'Agent Test',
                    10, 0.09, 5, 300, 275, 150, 125, 0)
        """, {'luna': luna})
    # An active in-transit order for the same SKU — must NOT reduce zile_stoc
    conn.execute("""
        INSERT INTO comenzi_furnizori (nr_comanda, furnizor, status, data_estimata_livrare)
        VALUES ('CMD-B2-1', 'TestBrandB2', 'confirmata', '2026-08-01')
    """)
    cid = conn.execute("SELECT id FROM comenzi_furnizori WHERE nr_comanda='CMD-B2-1'").fetchone()[0]
    conn.execute("""
        INSERT INTO comenzi_furnizori_linii (comanda_id, sku, cantitate_comandata)
        VALUES (?, 'SKU-B2-001', 30)
    """, (cid,))
    conn.commit()
    conn.close()

    import queries
    rows = queries.forecast_stoc_extended(furnizor='TestBrandB2')
    assert len(rows) == 1
    r = rows[0]
    # 30 avg/month sales -> daily rate 1/day -> stoc-only zile_stoc should be ~30 (30 buc / 1 buc/day)
    # If transit (30 more) were included, available=60 -> zile_stoc would be ~60
    assert r['zile_stoc'] < 45, (
        f"zile_stoc={r['zile_stoc']} appears to include the 30-unit in-transit order "
        "(physical stock alone should give ~30 days, not ~60)"
    )


def test_transit_eta_prefers_eta_column(db_path, client):
    conn = _conn(db_path)
    snap = _next_snapshot(conn)
    conn.execute("""
        INSERT INTO stoc (data_snapshot, cod_produs, cod_mare, sku, furnizor, gama,
                           cantitate, pret_achizitie, data_intrare)
        VALUES (?, 'B6-001', 'B6-001', 'SKU-B6-001', 'TestBrandB6', 'Ceai',
                5, 10.0, '2026-06-01')
    """, (snap,))
    conn.execute("""
        INSERT INTO comenzi_furnizori (nr_comanda, furnizor, status, data_estimata_livrare, eta)
        VALUES ('CMD-B6-1', 'TestBrandB6', 'confirmata', '2026-06-02', '2026-07-21')
    """)
    cid = conn.execute("SELECT id FROM comenzi_furnizori WHERE nr_comanda='CMD-B6-1'").fetchone()[0]
    conn.execute("""
        INSERT INTO comenzi_furnizori_linii (comanda_id, sku, cantitate_comandata)
        VALUES (?, 'SKU-B6-001', 10)
    """, (cid,))
    conn.commit()
    conn.close()

    import queries
    rows = queries.forecast_stoc_extended(furnizor='TestBrandB6')
    assert len(rows) == 1
    assert rows[0]['in_tranzit'][0]['eta'] == '2026-07-21', (
        "should prefer the newer `eta` column over the stale `data_estimata_livrare`"
    )


def test_monthly_sales_by_sku_survives_quote_in_export_code(db_path, client):
    conn = _conn(db_path)
    # A client code containing a single quote — breaks the old f-string-interpolated SQL
    conn.execute("""
        INSERT INTO clienti_export (tara_id, cod_client, nume_client, activ)
        VALUES (1, "O'BRIEN", 'OBrien Ltd', 1)
    """)
    conn.execute("""
        INSERT INTO tranzactii (luna, an, data_dl, sku, furnizor, cantitate,
                                 cod_produs, client, cod_client, agent,
                                 pret_vanzare, tva_pct, pret_cumparare,
                                 val_bruta, val_neta, val_achizitie, marja_bruta, discount_pct)
        VALUES (1, 2025, '2025-01-10', 'SKU-C3-001', 'TestBrandC3', 20,
                'C3-001', 'OBrien Ltd', "O'BRIEN", 'Agent Test',
                10, 0.09, 5, 200, 180, 100, 80, 0)
    """)
    conn.commit()
    conn.close()

    from forecast import forecast_logic
    result = forecast_logic._monthly_sales_by_sku('TestBrandC3')  # must not raise
    assert 'SKU-C3-001' in result
    assert result['SKU-C3-001']['export'].get(1) == 20, "sale should be attributed to export (O'BRIEN is active)"
    assert result['SKU-C3-001']['ro'].get(1, 0) == 0


def test_forecast_stoc_extended_velocity_mode(db_path, client):
    conn = _conn(db_path)
    snap = _next_snapshot(conn)
    conn.execute("""
        INSERT INTO stoc (data_snapshot, cod_produs, cod_mare, sku, furnizor, gama,
                           cantitate, pret_achizitie, data_intrare)
        VALUES (?, 'VEL-1', 'VEL-1', 'SKU-VEL-1', 'TestBrandVEL', 'Ceai',
                100, 10.0, '2026-06-01')
    """, (snap,))

    def _sale(data_expr, an, luna, qty, params=()):
        conn.execute(f"""
            INSERT INTO tranzactii (luna, an, data_dl, sku, furnizor, cantitate,
                                     cod_produs, client, cod_client, agent,
                                     pret_vanzare, tva_pct, pret_cumparare,
                                     val_bruta, val_neta, val_achizitie, marja_bruta, discount_pct)
            VALUES (:luna, :an, {data_expr}, 'SKU-VEL-1', 'TestBrandVEL', :qty,
                    'VEL-1', 'Client Test', 'C001', 'Agent Test',
                    10, 0.09, 5, 300, 275, 150, 125, 0)
        """, {'luna': luna, 'an': an, 'qty': qty, **dict(params)})

    # Recent sales (last 90 days): 90 units total -> 90-day velocity = 30/month
    for off in ('-10 days', '-40 days', '-70 days'):
        _sale("date('now', :off)", 2026, 1, 30, (('off', off),))
    # Older sales within the 3-year window but outside 90 days -> lift the 3-year avg
    for an in (2024, 2025):
        for luna in range(1, 13):
            _sale("printf('%04d-%02d-15', :an, :luna)", an, luna, 50)
    conn.commit()
    conn.close()

    import queries
    r90 = {r['sku']: r for r in
           queries.forecast_stoc_extended(furnizor='TestBrandVEL', vel='90zile')}['SKU-VEL-1']
    r3 = {r['sku']: r for r in
          queries.forecast_stoc_extended(furnizor='TestBrandVEL', vel='3ani')}['SKU-VEL-1']

    assert r90['vanzari_luna_avg'] == 30.0, "90zile mode must show the raw 90-day velocity"
    assert r3['vanzari_luna_avg'] != r90['vanzari_luna_avg'], (
        "3ani mode must show the seasonal 3-year velocity, different from 90-day"
    )
    # zile_stoc follows the selected velocity (same 100-unit stock, different divisor)
    assert r90['zile_stoc'] != r3['zile_stoc']
    # Default (no vel arg) preserves the historical 3-year behaviour
    rdef = {r['sku']: r for r in
            queries.forecast_stoc_extended(furnizor='TestBrandVEL')}['SKU-VEL-1']
    assert rdef['vanzari_luna_avg'] == r3['vanzari_luna_avg']


def test_forecast_export_honors_velocity_mode(db_path, client):
    import io
    conn = _conn(db_path)
    snap = _next_snapshot(conn)
    conn.execute("""
        INSERT INTO stoc (data_snapshot, cod_produs, cod_mare, sku, furnizor, gama,
                           cantitate, pret_achizitie, data_intrare)
        VALUES (?, 'VELX-1', 'VELX-1', 'SKU-VELX-1', 'TestBrandVELX', 'Ceai',
                100, 10.0, '2026-06-01')
    """, (snap,))
    for off in ('-10 days', '-40 days', '-70 days'):
        conn.execute("""
            INSERT INTO tranzactii (luna, an, data_dl, sku, furnizor, cantitate,
                                     cod_produs, client, cod_client, agent,
                                     pret_vanzare, tva_pct, pret_cumparare,
                                     val_bruta, val_neta, val_achizitie, marja_bruta, discount_pct)
            VALUES (1, 2026, date('now', :off), 'SKU-VELX-1', 'TestBrandVELX', 30,
                    'VELX-1', 'Client Test', 'C001', 'Agent Test',
                    10, 0.09, 5, 300, 275, 150, 125, 0)
        """, {'off': off})
    conn.commit()
    conn.close()

    import openpyxl
    resp = client.get('/export/forecast?brand=TestBrandVELX&vel=90zile')
    assert resp.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(resp.data), data_only=True).active
    headers = [str(c.value or '') for c in next(ws.iter_rows(min_row=1, max_row=1))]
    vcol = next(i for i, h in enumerate(headers) if h.startswith('Vânz./lună'))
    assert '90 zile' in headers[vcol], f"export header should name the mode: {headers[vcol]}"
    # find our SKU row and check the velocity value = 90/3 = 30
    sku_col = headers.index('SKU')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[sku_col] == 'SKU-VELX-1':
            assert row[vcol] == 30, f"90-day velocity should be 30, got {row[vcol]}"
            break
    else:
        raise AssertionError('SKU-VELX-1 not found in export')


def test_listing_changes_keys_are_normalized(db_path, client):
    conn = _conn(db_path)
    # ERP-style bare-EAN SKU (no parens) — ERP sometimes exports it this way
    bare_sku = 'PRODUS TEST C5 1234567890123'
    conn.execute("""
        INSERT INTO tranzactii (luna, an, data_dl, sku, furnizor, cantitate,
                                 cod_produs, client, cod_client, agent,
                                 pret_vanzare, tva_pct, pret_cumparare,
                                 val_bruta, val_neta, val_achizitie, marja_bruta, discount_pct)
        VALUES (7, 2026, date('now', '-10 days'), :sku, 'TestBrandC5', 5,
                'C5-001', 'New Client', 'NEWC5', 'Agent Test',
                10, 0.09, 5, 50, 45, 25, 20, 0)
    """, {'sku': bare_sku})
    conn.commit()
    conn.close()

    from forecast import forecast_logic
    changes = forecast_logic._listing_changes('TestBrandC5')
    normalized = forecast_logic._normalize_sku(bare_sku)
    assert normalized in changes, (
        f"expected normalized key {normalized!r} in {list(changes.keys())} "
        "so build_suggestion's normalized-key lookup finds it"
    )
