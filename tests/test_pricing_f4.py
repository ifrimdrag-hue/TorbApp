"""Pricing F4: prospect clients, potential articles, supplier offer import,
product photos."""
from io import BytesIO
import os
import sqlite3

import openpyxl
import pytest


def _oferta_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Code', 'Product name', 'Price EUR', 'EAN', 'Pcs/box'])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_prospect_create_and_reuse(client, db_path):
    rv = client.post('/api/preturi/clienti-prospect', json={'nume': 'Retail Nou SRL'})
    d = rv.get_json()
    assert d['ok'] and d['cod_client'].startswith('PROSPECT-')
    # same name -> same code, no duplicate
    rv2 = client.post('/api/preturi/clienti-prospect', json={'nume': 'retail nou srl'})
    assert rv2.get_json()['cod_client'] == d['cod_client']
    rv3 = client.post('/api/preturi/clienti-prospect', json={'nume': '  '})
    assert rv3.status_code == 400
    # prospect appears in the simulator client dropdown
    rv4 = client.get('/preturi/simulator?an=2026')
    assert b'Retail Nou SRL [prospect]' in rv4.data


def test_prospect_proposal_and_generic_listing(client, db_path):
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        INSERT OR IGNORE INTO produse (sku, descriere, furnizor, gama, activ)
        VALUES ('F4-1', 'Ceai F4', 'Basilur', 'Basilur', 1);
        INSERT OR IGNORE INTO costuri_landing (an, sku, moneda,
            pret_achizitie_valuta, curs_ron, pret_achizitie_ron, transport_pct,
            taxa_vamala_pct, alte_costuri_ron, landing_cost_ron)
        VALUES (2026, 'F4-1', 'USD', 10, 4.6, 46, 0, 0, 0, 46);
    """)
    conn.commit()
    conn.close()
    cod = client.post('/api/preturi/clienti-prospect',
                      json={'nume': 'Prospect Oferta SRL'}).get_json()['cod_client']
    rv = client.post('/api/preturi/propuneri', json={
        'an': 2026, 'cod_client': cod, 'titlu': 'Oferta prospect',
        'linii': [{'sku': 'F4-1', 'pret_propus': 69}]})
    pid = rv.get_json()['id']
    assert rv.get_json()['ok']
    # listing falls back to the prospect's generic template with its name
    rv = client.get(f'/preturi/propuneri/{pid}/listare.xlsx')
    ws = openpyxl.load_workbook(BytesIO(rv.data)).active
    assert 'Prospect Oferta SRL' in ws.cell(1, 1).value
    client.delete(f'/api/preturi/propuneri/{pid}')


def test_import_oferta_preview_and_import(client, db_path):
    buf = _oferta_xlsx([['NS-100', 'Biscuiti ovaz', 1.25, '590123', 12],
                        ['NS-101', 'Biscuiti cacao', 1.4, '590124', 12],
                        ['', 'rand invalid fara cod', 9, '', '']])
    rv = client.post('/api/preturi/import-oferta',
                     data={'file': (buf, 'oferta.xlsx'), 'actiune': 'preview'},
                     content_type='multipart/form-data')
    d = rv.get_json()
    assert d['ok'] and d['total_rows'] == 4 and d['rows'][0][0] == 'Code'

    buf = _oferta_xlsx([['NS-100', 'Biscuiti ovaz', 1.25, '590123', 12],
                        ['NS-101', 'Biscuiti cacao', 1.4, '590124', 12]])
    rv = client.post('/api/preturi/import-oferta', data={
        'file': (buf, 'oferta.xlsx'), 'actiune': 'import',
        'col_cod': 'A', 'col_denumire': 'B', 'col_pret': 'C',
        'col_ean': 'D', 'col_buc_bax': 'E', 'rand_start': '2',
        'furnizor': 'NewSnacks', 'moneda': 'EUR', 'curs': '5.0',
        'transport_pct': '10', 'taxa_vamala_pct': '0', 'an': '2026',
    }, content_type='multipart/form-data')
    d = rv.get_json()
    assert d['ok'] and d['create'] == 2 and d['nr_sarite'] == 0
    conn = sqlite3.connect(db_path)
    row = conn.execute("SELECT potential, furnizor, ean FROM produse "
                       "WHERE sku='NS-100'").fetchone()
    assert row == (1, 'NewSnacks', '590123')
    landing = conn.execute("SELECT landing_cost_ron FROM costuri_landing "
                           "WHERE sku='NS-100' AND an=2026").fetchone()[0]
    assert landing == pytest.approx(1.25 * 5.0 * 1.1)
    conn.close()
    # re-import: existing skus are skipped, not overwritten
    buf = _oferta_xlsx([['NS-100', 'Biscuiti ovaz', 2.0, '', '']])
    rv = client.post('/api/preturi/import-oferta', data={
        'file': (buf, 'oferta.xlsx'), 'actiune': 'import',
        'col_cod': 'A', 'col_denumire': 'B', 'col_pret': 'C',
        'rand_start': '2', 'furnizor': 'NewSnacks', 'moneda': 'EUR',
        'curs': '5.0', 'an': '2026',
    }, content_type='multipart/form-data')
    d = rv.get_json()
    assert d['create'] == 0 and d['nr_sarite'] == 1


def test_poza_upload_and_url(client, db_path, tmp_path):
    from PIL import Image
    conn = sqlite3.connect(db_path)
    conn.execute("INSERT OR IGNORE INTO produse (sku, descriere, furnizor, activ)"
                 " VALUES ('F4-P', 'Cu poza', 'Basilur', 1)")
    conn.commit()
    conn.close()
    img = BytesIO()
    Image.new('RGB', (50, 50), (0, 128, 0)).save(img, format='PNG')
    img.seek(0)
    rv = client.post('/api/preturi/poza/F4-P',
                     data={'file': (img, 'poza.png')},
                     content_type='multipart/form-data')
    d = rv.get_json()
    assert d['ok'] and d['src'] == '/static/product_images/F4-P.png'
    conn = sqlite3.connect(db_path)
    p = conn.execute("SELECT path FROM produse_media WHERE sku='F4-P' AND "
                     "principala=1").fetchone()[0]
    conn.close()
    assert p == 'app/static/product_images/F4-P.png'
    import paths
    saved = os.path.join(paths.BASE_DIR, p)
    assert os.path.exists(saved)
    os.unlink(saved)
    # bad extension rejected
    bad = BytesIO(b'x')
    rv = client.post('/api/preturi/poza/F4-P', data={'file': (bad, 'x.exe')},
                     content_type='multipart/form-data')
    assert rv.status_code == 400
