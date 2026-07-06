"""Pricing F4/F5: article-creation sheets and supplier price updates."""
from io import BytesIO
import sqlite3

import openpyxl
import pytest


@pytest.fixture()
def setup_f5(client, db_path):
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        INSERT OR IGNORE INTO produse (sku, descriere, furnizor, gama, brand,
            gramaj, ean, tva_pct, buc_cutie, hs_code, tara_origine, activ)
        VALUES ('F5-1', 'Ceai F5', 'Basilur', 'Basilur', 'Basilur',
                100, '4790000000001', 0.09, 24, '09023000', 'Sri Lanka', 1);
        INSERT OR IGNORE INTO costuri_landing (an, sku, moneda,
            pret_achizitie_valuta, curs_ron, pret_achizitie_ron, transport_pct,
            taxa_vamala_pct, alte_costuri_ron, landing_cost_ron)
        VALUES (2026, 'F5-1', 'USD', 10, 4.6, 46, 10, 0, 0, 50.6);
        INSERT OR IGNORE INTO produse_logistica (sku, unit_net_kg,
            unit_gross_kg, bax_l_mm, bax_w_mm, bax_h_mm, bax_gross_kg,
            bax_cbm, buc_bax, bax_palet, valabilitate_luni)
        VALUES ('F5-1', 0.1, 0.13, 400, 300, 250, 3.5, 0.03, 24, 96, 24);
        INSERT INTO tranzactii (an, cod_client, client, furnizor, sku)
        SELECT 2026, 'CL-F5', 'AUCHAN TEST SA', 'Basilur', 'F5-1'
        WHERE NOT EXISTS (SELECT 1 FROM tranzactii WHERE cod_client='CL-F5');
        INSERT INTO comenzi_furnizori (nr_comanda, furnizor, data_comanda)
        SELECT 'CMD-F5', 'Basilur', '2026-05-01'
        WHERE NOT EXISTS (SELECT 1 FROM comenzi_furnizori WHERE nr_comanda='CMD-F5');
        INSERT OR IGNORE INTO propuneri_pret (id, an, cod_client, titlu)
        VALUES (9005, 2026, 'CL-F5', 'F5 fisa');
        INSERT OR IGNORE INTO propuneri_pret_linii (propunere_id, sku,
            pret_actual, pret_propus) VALUES (9005, 'F5-1', 60, 69);
    """)
    conn.execute("""
        INSERT INTO comenzi_furnizori_linii (comanda_id, sku, pret_valuta,
            cod_furnizor)
        SELECT id, 'F5-1', 10.8, 'BAS-F5' FROM comenzi_furnizori
        WHERE nr_comanda='CMD-F5'
        AND NOT EXISTS (SELECT 1 FROM comenzi_furnizori_linii WHERE sku='F5-1')
    """)
    conn.commit()
    conn.close()
    yield
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        DELETE FROM propuneri_pret_linii WHERE propunere_id=9005;
        DELETE FROM propuneri_pret WHERE id=9005;
    """)
    conn.commit()
    conn.close()


def _sheet(resp):
    assert resp.status_code == 200
    return openpyxl.load_workbook(BytesIO(resp.data)).active


def test_fisa_generic(client, setup_f5):
    ws = _sheet(client.get('/preturi/propuneri/9005/fisa.xlsx'))
    assert ws.cell(4, 1).value == 'Cod articol'
    row = {ws.cell(4, c).value: ws.cell(5, c).value
           for c in range(1, 18)}
    assert row['Cod articol'] == 'F5-1'
    assert row['Buc/bax'] == 24 and row['Bax/palet'] == 96
    assert row['Bax LxlxH (mm)'] == '400x300x250'
    assert row['Pret propus fara TVA'] == 69


def test_fisa_auchan_template(client, setup_f5):
    ws = _sheet(client.get(
        '/preturi/propuneri/9005/fisa.xlsx?template=auchan_creare&valabil=01.08.2026'))
    hdr = {ws.cell(4, c).value: c for c in range(1, 24)}
    assert 'Cod Tarifar' in hdr and 'PCB (buc/bax)' in hdr
    assert ws.cell(5, hdr['Cod Tarifar']).value == '09023000'
    assert ws.cell(5, hdr['Denumire furnizor']).value == 'TORB LOGISTIC SRL'
    assert ws.cell(5, hdr['Bax Lungime (m)']).value == 0.4
    assert ws.cell(5, hdr['Ingrediente']).value is None  # manual field


def _lista_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Code', 'Price'])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_actualizare_diff_and_apply(client, setup_f5, db_path):
    # supplier code BAS-F5 resolves via the last order's cod_furnizor
    rv = client.post('/api/preturi/actualizare-preturi', data={
        'file': (_lista_xlsx([['BAS-F5', 11.5], ['XX-404', 9]]), 'lista.xlsx'),
        'actiune': 'diff', 'an': '2026', 'furnizor': 'Basilur',
        'col_cod': 'A', 'col_pret': 'B', 'rand_start': '2',
    }, content_type='multipart/form-data')
    d = rv.get_json()
    assert d['ok'] and d['nr_necunoscute'] == 1
    linie = d['diff'][0]
    assert linie['sku'] == 'F5-1' and linie['pret_vechi'] == 10
    assert linie['delta_pct'] == 15.0
    assert linie['pret_ultima_comanda'] == 10.8

    rv = client.post('/api/preturi/actualizare-preturi?an=2026&furnizor=Basilur',
                     json={'linii': [{'sku': 'F5-1', 'pret_nou': 11.5}]})
    d = rv.get_json()
    assert d['ok'] and d['aplicate'] == 1
    assert len(d['alerte']) == 1 and 'ultima comanda' in d['alerte'][0]
    conn = sqlite3.connect(db_path)
    row = conn.execute(
        "SELECT pret_achizitie_valuta, landing_cost_ron FROM costuri_landing "
        "WHERE sku='F5-1' AND an=2026").fetchone()
    conn.close()
    # landing recomputed with the kept 10% transport: 11.5*4.6*1.1
    assert row[0] == 11.5
    assert row[1] == pytest.approx(11.5 * 4.6 * 1.1, abs=0.01)


def test_actualizare_requires_known_supplier(client, setup_f5):
    rv = client.post('/api/preturi/actualizare-preturi', data={
        'file': (_lista_xlsx([['X', 1]]), 'l.xlsx'), 'actiune': 'diff',
        'an': '2026', 'furnizor': 'FurnizorInexistent',
        'col_cod': 'A', 'col_pret': 'B',
    }, content_type='multipart/form-data')
    assert rv.status_code == 400
