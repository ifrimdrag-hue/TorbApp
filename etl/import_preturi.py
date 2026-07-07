"""
Import SKUs and prices from Excel files into the pricing module tables.
Sources:
  - monitorizare_preturi_torb_actualizat.xlsx  → produse + costuri_landing
  - Oferta produse TORB LOGISTIC CU ORGANSIA 01.01.2026.xlsx → hs_codes + selling prices
  - lista pret Online_06.03.2026.xlsx          → current selling prices (primary)
  - PRETURI ACHIZITIE Delaviuda.xlsx           → EUR purchase prices
  - PRETURI ACHIZITIE CELMAR 2026.xlsx         → PLN purchase prices
"""
import sqlite3
import openpyxl
import os
import re

DB = 'data/torb.db'
DOCS = os.path.join('docs_input')

# ── EU TARIC customs duty rates (effective rate under GSP+) ─────────────────
HS_DUTY = {
    # Chapter 09 – pure tea: GSP+ Sri Lanka = 0%
    '09021000': {'mfn': 3.2, 'gsp': 0.0, 'desc': 'Ceai verde pliculete ≤3kg'},
    '09022000': {'mfn': 3.2, 'gsp': 0.0, 'desc': 'Ceai verde vrac'},
    '09023000': {'mfn': 0.0, 'gsp': 0.0, 'desc': 'Ceai negru pliculete ≤3kg'},
    '09024000': {'mfn': 0.0, 'gsp': 0.0, 'desc': 'Ceai negru vrac'},
    # Chapter 21 – preparate: GSP+ reduce dar NU elimina taxa (ceaiuri cu fructe, extracte)
    '21012000': {'mfn': 3.8, 'gsp': 2.2, 'desc': 'Extract/concentrat ceai'},
    '21069020': {'mfn': 9.0, 'gsp': 6.3, 'desc': 'Preparate fara grasimi lactate (ceai cu fructe)'},
    '21069098': {'mfn': 9.0, 'gsp': 6.3, 'desc': 'Preparate alimentare nca (ceai cu fructe/zahar)'},
    # Chapter 22 – bauturi cu continut de fruct/ceai: taxa MFN aplicabila
    '22029900': {'mfn': 9.6, 'gsp': 9.6, 'desc': 'Alte bauturi nealcoolice'},
    '22029110': {'mfn': 9.6, 'gsp': 9.6, 'desc': 'Bauturi pe baza de ceai/cafea'},
    # Chapter 17/18 – dulciuri/ciocolata
    '17049099': {'mfn': 8.0, 'gsp': 5.6, 'desc': 'Produse zaharoase'},
    '18069000': {'mfn': 8.0, 'gsp': 5.6, 'desc': 'Ciocolata si preparate'},
}

# ── Origin / duty rules per supplier ────────────────────────────────────────
SUPPLIER_ORIGIN = {
    'Basilur':   ('Sri Lanka',  'import_extraeu'),
    'Tipson':    ('Sri Lanka',  'import_extraeu'),
    'KingsLeaf': ('Sri Lanka',  'import_extraeu'),
    'Organsia':  ('Sri Lanka',  'import_extraeu'),
    'Delaviuda': ('Spania',     'eu'),
    'Celmar':    ('Polonia',    'eu'),
    'Torras':    ('Spania',     'eu'),
    'Leonex':    ('Romania',    'intern'),
}

# Spreadsheet Brand-column spellings -> canonical tranzactii furnizor for the
# Basilur virtual sub-brands (see docs/BUSINESS_LOGIC.md §5).
VIRTUAL_BRAND_CANON = {
    'KINGSLEAF': 'KingsLeaf',
    'KINGS LEAF': 'KingsLeaf',
    'TIPSON': 'Tipson',
    'TIPSON TEA': 'Tipson',
    'ORGANSIA': 'Organsia',
}

# ── Helpers ──────────────────────────────────────────────────────────────────
def clean_hs(code):
    if not code:
        return None
    return re.sub(r'\D', '', str(code))[:8] or None

def wb_open(path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

def cell_val(cell):
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
    return v

# ── Step 1: Exchange rates ────────────────────────────────────────────────────
def import_rates(conn):
    rates = [
        (2026, 'EUR', 5.0918),
        (2026, 'USD', 4.3226),
        (2026, 'PLN', 1.2007),
        (2026, 'RON', 1.0000),
        (2025, 'EUR', 4.9700),
        (2025, 'USD', 4.5200),
        (2025, 'PLN', 1.1500),
        (2025, 'RON', 1.0000),
    ]
    for an, moneda, curs in rates:
        conn.execute("""
            INSERT OR REPLACE INTO rate_schimb (an, moneda, curs_ron)
            VALUES (?,?,?)
        """, (an, moneda, curs))
    print(f'  Cursuri: {len(rates)} inserturi')

# ── Step 2: Products from monitorizare ───────────────────────────────────────
def import_monitorizare(conn):
    path = os.path.join(DOCS, 'LISTE PRET ACHIITIE 2026',
                        'monitorizare_preturi_torb_actualizat.xlsx')
    wb = wb_open(path)
    ws = wb['Monitorizare']
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else '' for c in rows[0]]

    col = {h: i for i, h in enumerate(headers)}
    products = {}
    landing = {}

    for row in rows[1:]:
        sku = cell_val_idx(row, col.get('Cod produs'))
        if not sku:
            continue
        sku = str(sku).strip()
        furnizor = str(row[col['Furnizor']] or '').strip()
        brand    = str(row[col.get('Brand', col['Furnizor'])] or furnizor).strip()
        descriere= str(row[col['Produs']] or '').strip()
        _d = descriere.upper()
        if _d.startswith("ORGANSIA") or _d.startswith("B.ECO ORGANSIA"):
            furnizor = "Organsia"
            brand = "Organsia"
        elif brand.upper() in VIRTUAL_BRAND_CANON:
            # Virtual sub-brands of Basilur: the sheet's Furnizor column says
            # Basilur (the real supplier), but the Brand column identifies the
            # sub-brand — normalize to the canonical tranzactii furnizor so
            # the four brands stay separate everywhere.
            furnizor = brand = VIRTUAL_BRAND_CANON[brand.upper()]
        categorie= str(row[col.get('Categorie', 0)] or '').strip() if col.get('Categorie') is not None else ''
        moneda   = str(row[col['Monedă']] or 'USD').strip()
        pret_cur = to_float(row[col.get('Preț curent')])
        curs_ron = to_float(row[col.get('Curs RON')])
        pret_ron = to_float(row[col.get('Preț curent RON')])
        multiplu = to_float(row[col.get('Multiplu')]) or 1
        units    = to_float(row[col.get('Unități / multiplu')]) or 1

        origem, tip_origine = SUPPLIER_ORIGIN.get(furnizor, ('', 'import_extraeu'))

        products[sku] = {
            'sku': sku, 'descriere': descriere, 'furnizor': furnizor,
            'brand': brand, 'categorie': categorie,
            'buc_cutie': int(units * multiplu) if units and multiplu else None,
            'tva_pct': 0.09,
            'origine': tip_origine, 'tara_origine': origem,
            'activ': 1,
        }
        if pret_cur and curs_ron:
            pret_achizitie_ron = pret_ron or (pret_cur * curs_ron)
            transport_pct = 10.0
            landing_cost = pret_achizitie_ron * (1 + transport_pct / 100)
            landing[sku] = {
                'an': 2026, 'sku': sku, 'moneda': moneda,
                'pret_achizitie_valuta': pret_cur,
                'curs_ron': curs_ron,
                'pret_achizitie_ron': round(pret_achizitie_ron, 4),
                'transport_pct': transport_pct,
                'taxa_vamala_pct': 0.0,
                'alte_costuri_ron': 0.0,
                'landing_cost_ron': round(landing_cost, 4),
            }

    # Insert products
    for p in products.values():
        conn.execute("""
            INSERT OR REPLACE INTO produse
                (sku, descriere, furnizor, brand, categorie, buc_cutie,
                 tva_pct, origine, tara_origine, activ)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (p['sku'], p['descriere'], p['furnizor'], p['brand'],
              p['categorie'], p['buc_cutie'], p['tva_pct'],
              p['origine'], p['tara_origine'], p['activ']))

    # Insert landing costs
    for lc in landing.values():
        conn.execute("""
            INSERT OR REPLACE INTO costuri_landing
                (an, sku, moneda, pret_achizitie_valuta, curs_ron,
                 pret_achizitie_ron, transport_pct, taxa_vamala_pct,
                 alte_costuri_ron, landing_cost_ron)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (lc['an'], lc['sku'], lc['moneda'], lc['pret_achizitie_valuta'],
              lc['curs_ron'], lc['pret_achizitie_ron'], lc['transport_pct'],
              lc['taxa_vamala_pct'], lc['alte_costuri_ron'], lc['landing_cost_ron']))

    print(f'  Produse: {len(products)} inserturi')
    print(f'  Costuri landing: {len(landing)} inserturi')
    wb.close()

# ── Step 3: HS codes + selling prices from Oferta ───────────────────────────
def import_oferta(conn, path, an=2026):
    wb = wb_open(path)
    count_hs, count_pv = 0, 0
    for sheet_name in wb.sheetnames:
        if 'instructi' in sheet_name.lower() or 'rezumat' in sheet_name.lower():
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row (look for 'Cod articol' or 'Cod EAN')
        hdr_row = None
        for i, row in enumerate(rows[:20]):
            if row and any(str(c or '').strip().lower() in
                          ('cod articol', 'cod ean', 'cod vamal') for c in row):
                hdr_row = i
                break
        if hdr_row is None:
            continue

        headers = [str(c or '').strip().lower() for c in rows[hdr_row]]
        def ci(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n.lower() in h:
                        return i
            return None

        col_sku   = ci(['cod articol'])
        col_pret  = ci(['pret unitar', 'pret unit'])
        col_tva   = ci(['cota tva', 'tva'])
        col_hs    = ci(['cod vamal', 'vamal', 'hs'])
        col_buc   = ci(['buc./box', 'buc/box', 'unitati'])
        col_ean   = ci(['cod ean', 'ean'])
        col_gram  = ci(['gramaj'])

        if col_sku is None:
            continue

        for row in rows[hdr_row + 1:]:
            if not row or not row[col_sku]:
                continue
            sku = str(row[col_sku]).strip()
            if not sku or sku.lower() in ('cod articol', 'total'):
                continue

            hs = clean_hs(row[col_hs]) if col_hs is not None else None
            if hs:
                duty_info = HS_DUTY.get(hs, {})
                tva = to_float(row[col_tva]) if col_tva is not None else None
                buc = to_int(row[col_buc]) if col_buc is not None else None
                gram = to_float(row[col_gram]) if col_gram is not None else None
                ean = str(row[col_ean]).strip() if col_ean is not None and row[col_ean] else None

                gsp_rate = duty_info.get('gsp', duty_info.get('mfn', 0))
                conn.execute("""
                    UPDATE produse SET
                        hs_code = COALESCE(hs_code, ?),
                        taxa_vamala_mfn_pct = ?,
                        taxa_vamala_pct = ?,
                        tva_pct = COALESCE(CASE WHEN ? IS NOT NULL THEN ? ELSE NULL END, tva_pct),
                        buc_cutie = COALESCE(CASE WHEN ? IS NOT NULL THEN ? ELSE NULL END, buc_cutie),
                        gramaj = COALESCE(CASE WHEN ? IS NOT NULL THEN ? ELSE NULL END, gramaj),
                        ean = COALESCE(CASE WHEN ? IS NOT NULL THEN ? ELSE NULL END, ean)
                    WHERE sku = ?
                """, (hs, duty_info.get('mfn', 0), gsp_rate,
                      tva, tva,
                      buc, buc,
                      gram, gram,
                      ean, ean,
                      sku))
                count_hs += 1

            pret = to_float(row[col_pret]) if col_pret is not None else None
            if pret and pret > 0:
                conn.execute("""
                    INSERT OR REPLACE INTO preturi_vanzare
                        (an, sku, cod_client, pret_vanzare_ron, activ)
                    VALUES (?, ?, NULL, ?, 1)
                """, (an, sku, pret))
                count_pv += 1

    wb.close()
    print(f'  HS codes actualizate: {count_hs}')
    print(f'  Preturi vanzare standard: {count_pv}')

# ── Step 4: Delaviuda EUR prices ─────────────────────────────────────────────
def import_delaviuda(conn):
    path = os.path.join(DOCS, 'LISTE PRET ACHIITIE 2026',
                        'PRETURI ACHIZITIE  Delaviuda.xlsx')
    wb = wb_open(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    hdr = None
    for i, row in enumerate(rows[:15]):
        if row and any(str(c or '').upper() in ('INTERNAL CODE', 'PRODUCT') for c in row):
            hdr = i
            break
    if hdr is None:
        wb.close()
        return

    headers = [str(c or '').strip().upper() for c in rows[hdr]]
    def ci(name):
        for i, h in enumerate(headers):
            if name.upper() in h:
                return i
        return None

    col_sku  = ci('INTERNAL CODE')
    col_pret = ci('UNIT PRICE')
    col_buc  = ci('UNITS/CASE')
    curs_eur = 5.0918
    count = 0

    for row in rows[hdr + 1:]:
        if not row or not (col_sku is not None and row[col_sku]):
            continue
        sku = str(row[col_sku]).strip()
        pret_eur = to_float(row[col_pret]) if col_pret is not None else None
        buc = to_int(row[col_buc]) if col_buc is not None else None

        if not sku or not pret_eur:
            continue

        pret_ron = pret_eur * curs_eur
        landing  = pret_ron * 1.10

        conn.execute("""
            INSERT OR REPLACE INTO costuri_landing
                (an, sku, moneda, pret_achizitie_valuta, curs_ron,
                 pret_achizitie_ron, transport_pct, taxa_vamala_pct,
                 alte_costuri_ron, landing_cost_ron)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (2026, sku, 'EUR', pret_eur, curs_eur,
              round(pret_ron, 4), 10.0, 0.0, 0.0, round(landing, 4)))

        if buc:
            conn.execute("UPDATE produse SET buc_cutie=? WHERE sku=? AND buc_cutie IS NULL",
                         (buc, sku))
        count += 1

    wb.close()
    print(f'  Delaviuda: {count} costuri landing EUR')

# ── Step 5a: Celmar PLN prices (using EAN from tranzactii as SKU) ─────────────
def import_celmar(conn):
    # Build EAN→(descriere, ean_str) map from tranzactii for Celmar products
    # Format in tranzactii.sku: "CELMAR MENTA 5947226224061"
    celmar_tranz = {}
    for (sku_raw,) in conn.execute(
        "SELECT DISTINCT sku FROM tranzactii WHERE furnizor LIKE '%elmar%' AND sku IS NOT NULL"
    ).fetchall():
        m = re.search(r'\b(\d{13})\b', sku_raw or '')
        if m:
            ean = m.group(1)
            # Extract keyword: word(s) between 'CELMAR ' and the EAN
            kw_m = re.search(r'CELMAR\s+(.+?)\s+\d{13}', sku_raw.upper())
            keyword = kw_m.group(1).strip() if kw_m else ''
            if keyword:  # skip entries with no keyword (e.g. "CELMAR 5902795030932")
                celmar_tranz[keyword] = (sku_raw.strip(), ean)

    path = os.path.join(DOCS, 'LISTE PRET ACHIITIE 2026',
                        'PRETURI ACHIZITIE CELMAR 2026.xlsx')
    wb = wb_open(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    hdr = None
    for i, row in enumerate(rows[:15]):
        if row and any(str(c or '').upper() in ('PRODUCT', 'PRICE / PCS') for c in row):
            hdr = i
            break
    if hdr is None:
        wb.close()
        return

    headers = [str(c or '').strip().upper() for c in rows[hdr]]
    col_prod = next((i for i, h in enumerate(headers) if 'PRODUCT' in h), None)
    col_pret = next((i for i, h in enumerate(headers) if 'PRICE' in h), None)
    curs_pln = conn.execute(
        "SELECT curs_ron FROM rate_schimb WHERE an=2026 AND moneda='PLN'"
    ).fetchone()
    curs_pln = curs_pln[0] if curs_pln else 1.2007
    count = 0

    def extract_ro_keyword(prod_str):
        # Normalize non-breaking spaces and strip
        s = prod_str.replace('\xa0', ' ').strip()
        # Romanian keyword is the last uppercase word(s) after closing paren
        m = re.search(r'\)\s+([A-Z][A-Z ]+)$', s)
        if m:
            return m.group(1).strip()
        # Fallback: everything after last ')'
        if ')' in s:
            after = s.rsplit(')', 1)[-1].strip()
            if after:
                return after.upper()
        # Last resort: last word
        return s.split()[-1].upper() if s else ''

    for row in rows[hdr + 1:]:
        if not row or not (col_prod is not None and row[col_prod]):
            continue
        prod = str(row[col_prod]).strip()
        pret_pln = to_float(row[col_pret]) if col_pret is not None else None
        if not prod or not pret_pln:
            continue

        ro_kw = extract_ro_keyword(prod)
        # Try to find EAN from tranzactii by keyword match
        ean = None
        for kw, (desc, e) in celmar_tranz.items():
            if kw and ro_kw and (ro_kw in kw or kw in ro_kw):
                ean = e
                break

        if ean:
            sku = ean
            descriere = f'CELMAR {ro_kw}'
        else:
            sku = f'CELMAR-{ro_kw[:20].replace(" ","_")}'
            descriere = f'CELMAR {ro_kw}'

        # Insert/update produse
        conn.execute("""
            INSERT OR IGNORE INTO produse
                (sku, descriere, furnizor, brand, categorie, tva_pct, origine, tara_origine, activ)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (sku, descriere, 'Celmar', 'Celmar', 'Ceai', 0.09, 'import_extraeu', 'Polonia', 1))

        pret_ron = pret_pln * curs_pln
        landing  = pret_ron * 1.10
        conn.execute("""
            INSERT OR REPLACE INTO costuri_landing
                (an, sku, moneda, pret_achizitie_valuta, curs_ron,
                 pret_achizitie_ron, transport_pct, taxa_vamala_pct,
                 alte_costuri_ron, landing_cost_ron)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (2026, sku, 'PLN', pret_pln, curs_pln,
              round(pret_ron, 4), 10.0, 0.0, 0.0, round(landing, 4)))
        count += 1

    wb.close()
    print(f'  Celmar: {count} produse + costuri landing PLN')


# ── Step 5b: Leonex RON selling prices (EAN as SKU) ──────────────────────────
def import_leonex(conn):
    path = os.path.join(DOCS, 'Liste de preturi',
                        'Oferta de pret Leonex_05.03.2026.xlsx')
    if not os.path.exists(path):
        print('  Leonex: fisier lipsa')
        return
    wb = wb_open(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Header: 'Denumire produs', 'Buc/box', 'Pret fara TVA', 'Pret cu TVA', 'Cod ean'
    hdr = None
    for i, row in enumerate(rows[:25]):
        if row and any(str(c or '').lower().startswith('denumire') or
                       str(c or '').lower() == 'cod ean' for c in row):
            hdr = i
            break
    if hdr is None:
        wb.close()
        return

    headers = [str(c or '').strip().lower() for c in rows[hdr]]
    col_den  = next((i for i, h in enumerate(headers) if 'denumire' in h), None)
    col_buc  = next((i for i, h in enumerate(headers) if 'buc' in h), None)
    col_pret = next((i for i, h in enumerate(headers) if 'fara tva' in h or 'pret fara' in h), None)
    col_ean  = next((i for i, h in enumerate(headers) if 'ean' in h or 'cod ean' in h), None)

    count = 0
    for row in rows[hdr + 1:]:
        if not row:
            continue
        def gcell(col):
            return row[col] if col is not None and col < len(row) else None
        den  = str(gcell(col_den)).strip() if gcell(col_den) else ''
        ean_raw = gcell(col_ean)
        try:
            ean = str(int(float(ean_raw))) if ean_raw is not None else None
        except (ValueError, TypeError):
            ean = None
        pret = to_float(gcell(col_pret))
        buc  = to_int(gcell(col_buc))

        if not ean or not den or len(ean) < 8:
            continue

        sku = ean  # Use EAN as SKU — matches tranzactii format

        conn.execute("""
            INSERT OR IGNORE INTO produse
                (sku, descriere, furnizor, brand, categorie, buc_cutie,
                 tva_pct, origine, tara_origine, activ)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (sku, den, 'Leonex', 'Leonex', 'Igiena', buc,
              0.09, 'intern', 'Romania', 1))

        if pret and pret > 0:
            conn.execute("""
                INSERT OR REPLACE INTO preturi_vanzare
                    (an, sku, cod_client, pret_vanzare_ron, activ)
                VALUES (?,?,NULL,?,1)
            """, (2026, sku, round(pret, 4)))

        count += 1

    wb.close()
    print(f'  Leonex: {count} produse + preturi vanzare RON')

# ── Step 6: Update selling prices from most recent list ──────────────────────
def import_preturi_online(conn):
    path = os.path.join(DOCS, 'Liste de preturi',
                        'lista pret Online_06.03.2026.xlsx')
    wb = wb_open(path)
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        hdr = None
        for i, row in enumerate(rows[:20]):
            if row and any(str(c or '').strip().lower() in ('cod articol',) for c in row):
                hdr = i
                break
        if hdr is None:
            continue
        headers = [str(c or '').strip().lower() for c in rows[hdr]]
        col_sku  = next((i for i, h in enumerate(headers) if 'cod articol' in h), None)
        col_pret = next((i for i, h in enumerate(headers) if 'pret unitar' in h or 'pret unit' in h), None)
        if col_sku is None or col_pret is None:
            continue
        for row in rows[hdr + 1:]:
            if not row or not row[col_sku]:
                continue
            sku = str(row[col_sku]).strip()
            pret = to_float(row[col_pret])
            if not sku or not pret or pret <= 0:
                continue
            conn.execute("""
                INSERT OR REPLACE INTO preturi_vanzare
                    (an, sku, cod_client, pret_vanzare_ron, activ)
                VALUES (?,?,NULL,?,1)
            """, (2026, sku, pret))
            count += 1
    wb.close()
    print(f'  Preturi online actuale: {count} actualizari')

# ── Helpers ───────────────────────────────────────────────────────────────────
def cell_val_idx(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    return str(v).strip() if v is not None else None

def to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '.').replace(' ', ''))
    except Exception:
        return None

def to_int(v):
    f = to_float(v)
    return int(f) if f is not None else None

# ── Step 8: Import all remaining unmatched products from tranzactii ──────────
def import_unmatched_from_tranzactii(conn):
    """
    For every transaction SKU not yet in catalog, create a produse entry.
    Uses EAN (13-digit) as SKU when available, else 5-digit product code + '-00',
    else a slugified version of the description.
    """
    existing_skus = set(r[0] for r in conn.execute('SELECT sku FROM produse').fetchall())

    def find_match(raw):
        if not raw:
            return None
        if raw in existing_skus:
            return raw
        # 13-digit EAN anywhere
        m13 = re.search(r'\b(\d{13})\b', raw)
        if m13 and m13.group(1) in existing_skus:
            return m13.group(1)
        # Code before '(' — handles Toras pattern "DESC-529 (EAN)"
        mc = re.search(r'-(\d+)\s*\(', raw)
        if mc:
            for suffix in ('', '-00'):
                k = mc.group(1) + suffix
                if k in existing_skus:
                    return k
        # 5-digit standalone code — handles Basilur "71618-EAN"
        m5 = re.search(r'\b(\d{5})\b', raw)
        if m5:
            for suffix in ('', '-00'):
                k = m5.group(1) + suffix
                if k in existing_skus:
                    return k
        return None

    # Group unmatched transactions
    rows = conn.execute("""
        SELECT sku, furnizor, COUNT(*) as cnt, SUM(val_neta) as val,
               AVG(pret_vanzare) as avg_pret
        FROM tranzactii
        WHERE sku IS NOT NULL AND sku != ''
        GROUP BY sku
    """).fetchall()

    inserted = skipped = 0
    for row in rows:
        sku_raw = row[0]
        furnizor = (row[1] or '').strip()

        if find_match(sku_raw):
            skipped += 1
            continue

        # Determine best SKU key — prefer short code over long EAN
        mc  = re.search(r'-(\d+)\s*\(', sku_raw)   # "DESC-529 (EAN)"
        m13 = re.search(r'\b(\d{13})\b', sku_raw)
        m5  = re.search(r'\b(\d{5})\b', sku_raw)
        if mc and not (m5 and len(m5.group(1)) == 5 and m5.start() < mc.start()):
            # Short code before '(' — use as-is (no suffix, Toras style)
            sku = mc.group(1)
        elif m5:
            sku = m5.group(1) + '-00'
        elif m13:
            sku = m13.group(1)
        else:
            # Use raw transaction SKU directly so future lookups match exactly
            sku = sku_raw[:200]

        if sku in existing_skus:
            skipped += 1
            continue

        # Clean description: strip trailing EAN / codes
        descriere = re.sub(r'\s+\d{13}$', '', sku_raw).strip()
        descriere = re.sub(r'\s+\d{13}\b', '', descriere).strip()

        # Normalize furnizor name
        furn_norm = furnizor
        if furnizor.lower() in ('toras',):
            furn_norm = 'Toras'
        brand = furn_norm

        # Guess origin
        origine = 'import_extraeu'
        if furn_norm in ('Leonex', 'Colian', 'Cosmetice'):
            origine = 'intern'
        elif furn_norm in ('Torras', 'Toras', 'Delaviuda'):
            origine = 'eu'

        avg_pret = row[4]
        conn.execute("""
            INSERT OR IGNORE INTO produse
                (sku, descriere, furnizor, brand, tva_pct, origine, activ)
            VALUES (?,?,?,?,?,?,?)
        """, (sku, descriere[:200], furn_norm, brand, 0.09, origine, 1))

        # Add selling price if we have average transaction price
        if avg_pret and avg_pret > 0:
            conn.execute("""
                INSERT OR IGNORE INTO preturi_vanzare
                    (an, sku, cod_client, pret_vanzare_ron, activ)
                VALUES (2026,?,NULL,?,1)
            """, (sku, round(avg_pret, 4)))

        existing_skus.add(sku)
        inserted += 1

    print(f'  Import tranzactii: {inserted} produse noi, {skipped} deja existente')


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    conn = sqlite3.connect(DB, timeout=30)
    conn.row_factory = sqlite3.Row

    print('1. Cursuri valutare...')
    import_rates(conn)

    print('2. Produse din Monitorizare...')
    import_monitorizare(conn)

    print('3. Coduri HS + preturi vanzare (Oferta)...')
    import_oferta(conn, os.path.join(DOCS, 'Liste de preturi',
        'Oferta produse TORB LOGISTIC CU ORGANSIA 01.01.2026.xlsx'))

    print('4. Preturi achizitie Delaviuda (EUR)...')
    import_delaviuda(conn)

    print('5. Preturi achizitie Celmar (PLN) + produse...')
    import_celmar(conn)

    print('6. Produse + preturi vanzare Leonex (RON)...')
    import_leonex(conn)

    print('7. Preturi vanzare online (cele mai recente)...')
    import_preturi_online(conn)

    print('8. Import produse lipsa din tranzactii...')
    import_unmatched_from_tranzactii(conn)

    conn.commit()
    conn.close()

    # Summary
    conn2 = sqlite3.connect(DB)
    n_sku  = conn2.execute('SELECT COUNT(*) FROM produse').fetchone()[0]
    n_lc   = conn2.execute('SELECT COUNT(*) FROM costuri_landing').fetchone()[0]
    n_pv   = conn2.execute('SELECT COUNT(*) FROM preturi_vanzare').fetchone()[0]
    n_rs   = conn2.execute('SELECT COUNT(*) FROM rate_schimb').fetchone()[0]
    conn2.close()
    print('\n=== IMPORT FINALIZAT ===')
    print(f'  SKU-uri: {n_sku}')
    print(f'  Costuri landing: {n_lc}')
    print(f'  Preturi vanzare: {n_pv}')
    print(f'  Rate schimb: {n_rs}')
