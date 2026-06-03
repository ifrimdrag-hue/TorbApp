"""
Import vânzări din exportul direct ERP (format Sheet1, fără câmpuri pre-calculate).

Format fișier: docs_input/DD.MM.YYYY/vanzari DD.MM.YYYY.xlsx
Sheet: Sheet1
Câmpuri calculate: luna, an, val_bruta, val_neta, val_achizitie, marja_bruta, furnizor, sku

Usage:
    python import_vanzari_erp.py [<cale_fisier.xlsx>]
    # fără argumente: detectează automat cel mai recent folder datat
"""

import sys
import os
import re
import sqlite3
import openpyxl
from datetime import datetime, date

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

DB_PATH = "data/torb.db"
DOCS_PATH = "docs_input"

# Normalize known agent name variants to the canonical form used historically.
# ERP-ul a inceput sa exporte uppercase pentru unele nume — pastram forma istorica.
AGENT_NAME_MAP = {
    "OANA FILIP": "Oana Filip",
}

# Cod-uri client de excluse la import — facturate fictiv prin intermediar.
# TOBRA INVEST SRL (cod 719) este intermediar economic catre Auchan;
# vanzarile reale Tobra→Auchan sunt importate separat prin
# import_vanzari_tobra_auchan.py si atribuite agentului Oana.
SKIP_COD_CLIENT = {"719"}

DB_COLS = [
    "luna", "an", "data_dl", "nr_dl", "nr_factura", "nr_comanda",
    "cod_produs", "sku", "furnizor", "um",
    "cantitate", "pret_vanzare", "tva_pct", "pret_cumparare",
    "val_bruta", "val_neta", "val_achizitie", "val_usd", "marja_bruta",
    "discount_pct", "discount_val",
    "client", "cod_client", "cui_client", "tip_client",
    "oras_client", "judet_client", "adresa_client",
    "agent", "adr_livrare", "locatie",
]

# Maps ERP column name → canonical DB name
COL_MAP = {
    "datadl":       "data_dl",
    "nrdl":         "nr_dl",
    "cantit":       "cantitate",
    "pvanz":        "pret_vanzare",
    "tva":          "tva_pct",
    "pcump":        "pret_cumparare",
    "den_a":        "client",
    "factout":      "nr_factura",
    "numeag":       "agent",
    "procent":      "discount_pct",
    "adr_livr":     "adr_livrare",
    "nrcomandam":   "nr_comanda",
    "codprod":      "cod_produs",
    "den_b":        "sku",
    "discount":     "discount_val",
    "codcli":       "cod_client",
    "adresa":       "adresa_client",
    "discproc":     "discount_pct",   # fallback if procent=0
    "locatie":      "locatie",
    "numetipcli":   "tip_client",
    "cfcli":        "cui_client",
    "localcli":     "oras_client",
    "judet":        "judet_client",
    "um":           "um",
}


def _furnizor_from_prefix(sku: str) -> str | None:
    """Returns canonical furnizor based on SKU prefix, or None if ambiguous."""
    if not sku:
        return None
    s = str(sku).strip()
    if s.startswith("B.") or s.startswith('B."') or s.startswith("WB."):
        return "Basilur"
    if s.startswith("KL "):
        return "KingsLeaf"
    if s.upper().startswith("CELMAR") or s.startswith("C."):
        return "Celmar"
    if "5902795" in s or "5902480" in s:
        return "Celmar"
    if s.startswith("T."):
        return "Toras"
    if s.startswith("TS "):
        return "Tipson"
    if s.startswith("DEL.") or s.startswith("ALM."):
        return "Delaviuda"
    su = s.upper()
    for leonex_prefix in ("LEONEX", "BETISOARE", "DISCURI", "VATA ", "BILUTE", "SERVETELE",
                          "W.BETISOARE", "W.DISCURI", "W.SERVETELE"):
        if su.startswith(leonex_prefix):
            return "Leonex"
    for solvex_marker in ("MISS MAGIC", "SAMPON BLUE MAGIC", "STAND VOPSEA"):
        if solvex_marker in su:
            return "Solvex"
    if su.startswith("IMAJ "):
        return "Solvex"
    if su.startswith("HORECA ") or su.startswith("H "):
        return "Basilur"
    if (su.startswith("CUTIE HORECA") or su.startswith("CUTIE LEMN")
            or su.startswith("CUTIE INCHISA") or su.startswith("PUNGA ")
            or su.startswith("PUNGI ") or su.startswith("PAHAR ")):
        return "Basilur"
    cosm_markers = ("PRIME RENEWING", "INTENSE REGEN", "REGENERATING MASK",
                    "V-LIFT", "V-FIRM", "ELIXIR ", "VITAL ", "DETO2X",
                    "LUMI BOOST", "LUMIMASK", "H2O BOOST", "FLUID FALLS",
                    "BUBBLE FALLS", "ICY FALLS", "HYDRA3",
                    "MOISTURIZING ", "HUILE ", "LADY CODE", "BIO REV ",
                    "PURIFYING PACK", "SEA BLISS", "CREME DE MASQUE",
                    "HAND 24 HOUR", "PRIMARY VEIL", "BATHROBE ",
                    "HAIR BRUSH", "HEADBAND ")
    for m in cosm_markers:
        if su.startswith(m):
            return "Cosmetice"
    return None


def derive_furnizor(sku: str, cp_lookup: dict, cod_produs: str) -> str:
    """Determine furnizor: prefer SKU prefix when unambiguous, fall back to
    cod_produs lookup from history.

    Important: prefix takes priority over lookup. Otherwise, a single
    historical mis-assignment (e.g. a Celmar product imported once with
    furnizor='Basilur') keeps perpetuating itself on every subsequent import.
    """
    prefix_match = _furnizor_from_prefix(sku)
    if prefix_match is not None:
        return prefix_match
    # Ambiguous SKU (no clear prefix) → fall back to history lookup
    if cod_produs and cod_produs in cp_lookup:
        return cp_lookup[cod_produs]
    return "Altele"


def normalize_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val)


def normalize_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalize_num(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def find_latest_erp_file():
    """Find vanzari*.xlsx in the most recent dated folder (DD.MM.YYYY)."""
    date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    candidates = []
    for entry in os.listdir(DOCS_PATH):
        if date_pattern.match(entry):
            folder = os.path.join(DOCS_PATH, entry)
            if os.path.isdir(folder):
                for f in os.listdir(folder):
                    if f.lower().startswith("vanzari") and f.lower().endswith(".xlsx"):
                        day, month, year = entry.split(".")
                        folder_date = date(int(year), int(month), int(day))
                        candidates.append((folder_date, os.path.join(folder, f)))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def build_cod_furnizor_lookup(conn):
    """Build cod_produs → furnizor dict from existing tranzactii."""
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT cod_produs, furnizor FROM tranzactii WHERE furnizor IS NOT NULL")
    return {str(row[0]): row[1] for row in cursor.fetchall()}


def read_erp_sheet(filepath):
    print(f"  Citesc: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Try Sheet1, then first available sheet
    sheet_name = "Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    header = None
    rows_raw = []

    for raw_row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in raw_row):
            continue
        if header is None:
            header = [str(c).strip() if c is not None else "" for c in raw_row]
            continue
        rows_raw.append(dict(zip(header, raw_row)))

    wb.close()
    print(f"    → {len(rows_raw):,} rânduri brute citite")
    return rows_raw


def process_rows(rows_raw, cp_lookup):
    records = []
    skipped_intermediary = 0
    for raw in rows_raw:
        # Skip clienți intermediari (ex: TOBRA INVEST SRL) — facturate prin
        # alt flux și importate separat ca vânzări către clientul real.
        cod_cli_raw = raw.get("codcli")
        if cod_cli_raw is not None:
            try:
                cod_cli_str = str(int(float(cod_cli_raw)))
            except (ValueError, TypeError):
                cod_cli_str = str(cod_cli_raw).strip()
            if cod_cli_str in SKIP_COD_CLIENT:
                skipped_intermediary += 1
                continue

        # Map ERP column names to canonical names
        mapped = {}
        for erp_col, db_col in COL_MAP.items():
            val = raw.get(erp_col)
            if val is not None and db_col not in mapped:
                mapped[db_col] = val

        # Derived: luna, an from data_dl
        data_dl_raw = raw.get("datadl")
        if data_dl_raw is None:
            continue
        data_dl = normalize_date(data_dl_raw)
        if data_dl is None:
            continue

        if isinstance(data_dl_raw, (datetime, date)):
            dt = data_dl_raw if isinstance(data_dl_raw, date) else data_dl_raw.date()
        else:
            try:
                dt = datetime.strptime(data_dl[:10], "%Y-%m-%d").date()
            except ValueError:
                continue

        mapped["luna"] = dt.month
        mapped["an"] = dt.year
        mapped["data_dl"] = data_dl

        # Derived: furnizor from lookup or SKU prefix
        sku = normalize_str(raw.get("den_b"))
        cod_produs = normalize_str(raw.get("codprod"))
        mapped["furnizor"] = derive_furnizor(sku or "", cp_lookup, cod_produs or "")
        mapped["sku"] = sku

        # Calculated financials
        cantitate = normalize_num(raw.get("cantit")) or 0
        pvanz = normalize_num(raw.get("pvanz")) or 0
        pcump = normalize_num(raw.get("pcump")) or 0
        discount = normalize_num(raw.get("discount")) or 0

        val_bruta = round(cantitate * pvanz, 4)
        val_neta = round(val_bruta - discount, 4)
        val_achizitie = round(cantitate * pcump, 4)
        marja_bruta = round(val_neta - val_achizitie, 4)

        mapped["val_bruta"] = val_bruta
        mapped["val_neta"] = val_neta
        mapped["val_achizitie"] = val_achizitie
        mapped["marja_bruta"] = marja_bruta
        mapped["val_usd"] = None  # not available in direct ERP export

        # Type normalization
        record = {}
        for col in DB_COLS:
            val = mapped.get(col)
            if col == "data_dl":
                record[col] = normalize_date(val) if not isinstance(val, str) else val
            elif col in ("luna", "an"):
                record[col] = int(val) if val is not None else None
            elif col == "cod_client":
                try:
                    record[col] = int(float(val)) if val is not None else None
                except (ValueError, TypeError):
                    record[col] = normalize_str(val)
            elif col in (
                "cantitate", "pret_vanzare", "tva_pct", "pret_cumparare",
                "val_bruta", "val_neta", "val_achizitie", "val_usd",
                "marja_bruta", "discount_pct", "discount_val",
            ):
                record[col] = normalize_num(val)
            else:
                record[col] = normalize_str(val)

        # Normalize agent name to canonical historical form
        if record.get("agent") in AGENT_NAME_MAP:
            record["agent"] = AGENT_NAME_MAP[record["agent"]]

        records.append(record)

    if skipped_intermediary:
        print(f"    → {skipped_intermediary:,} rânduri sărite (clienți intermediari: {sorted(SKIP_COD_CLIENT)})")
    return records


def aggregate_records(records):
    """Agregă rânduri cu aceeași cheie (nr_dl, cod_produs, nr_factura).

    ERP-ul exportă câte un rând per magazin (lanțuri ca Kaufland, Carrefour),
    toate cu aceeași factură și produs. Fără locație distinctivă în export,
    singura diferență e cantitatea. Le sumăm pentru a obține totalul corect.
    """
    from collections import OrderedDict
    groups = OrderedDict()
    for r in records:
        # Include pret_vanzare in key so that free promo lines (pvanz=0) and
        # paid lines for the same product/invoice are never merged together.
        key = (r["nr_dl"], r["cod_produs"], r["nr_factura"], r["pret_vanzare"])
        if key not in groups:
            groups[key] = []
        groups[key].append(r)

    result = []
    aggregated_count = 0
    for key, group in groups.items():
        if len(group) == 1:
            result.append(group[0])
            continue
        aggregated_count += len(group) - 1
        base = dict(group[0])
        total_cantitate = sum(r["cantitate"] or 0 for r in group)
        total_discount = sum(r["discount_val"] or 0 for r in group)
        pvanz = base["pret_vanzare"] or 0
        pcump = base["pret_cumparare"] or 0
        total_val_bruta = round(total_cantitate * pvanz, 4)
        base["cantitate"] = total_cantitate
        base["discount_val"] = total_discount
        base["val_bruta"] = total_val_bruta
        base["val_neta"] = round(total_val_bruta - total_discount, 4)
        base["val_achizitie"] = round(total_cantitate * pcump, 4)
        base["marja_bruta"] = round(base["val_neta"] - base["val_achizitie"], 4)
        result.append(base)

    if aggregated_count:
        print(f"    → Agregare livrări multiple: {aggregated_count:,} rânduri cumulate în {aggregated_count} grupuri")
    return result


_CONFLICT_KEY = {"nr_dl", "cod_produs", "nr_factura", "pret_vanzare"}
_UPDATE_COLS = [c for c in DB_COLS if c not in _CONFLICT_KEY]


def insert_rows(conn, records):
    placeholders = ", ".join(["?" for _ in DB_COLS])
    col_names = ", ".join(DB_COLS)
    update_clause = ", ".join(f"{c} = excluded.{c}" for c in _UPDATE_COLS)
    sql = (
        f"INSERT INTO tranzactii ({col_names}) VALUES ({placeholders})"
        f" ON CONFLICT(nr_dl, cod_produs, nr_factura, pret_vanzare)"
        f" DO UPDATE SET {update_clause}"
    )
    data = [[r[c] for c in DB_COLS] for r in records]
    cursor = conn.cursor()
    cursor.executemany(sql, data)
    inserted = cursor.rowcount
    conn.commit()
    return inserted


def run(filepath=None):
    if filepath is None:
        filepath = find_latest_erp_file()
        if filepath is None:
            print("EROARE: Nu am găsit niciun fișier vanzari*.xlsx în foldere datate.")
            return 0

    conn = sqlite3.connect(DB_PATH)
    cp_lookup = build_cod_furnizor_lookup(conn)

    rows_raw = read_erp_sheet(filepath)
    records = process_rows(rows_raw, cp_lookup)
    print(f"    → {len(records):,} rânduri procesate")
    records = aggregate_records(records)
    print(f"    → {len(records):,} rânduri după agregare")

    inserted = insert_rows(conn, records)
    skipped = len(records) - inserted
    print(f"    → Inserate: {inserted:,} | Duplicate ignorate: {skipped:,}")

    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*), MIN(data_dl), MAX(data_dl) FROM tranzactii")
    total, d_min, d_max = cursor.fetchone()
    print(f"    → Total tranzactii în DB: {total:,} | {d_min} → {d_max}")

    conn.close()
    return inserted


if __name__ == "__main__":
    filepath = sys.argv[1] if len(sys.argv) > 1 else None
    run(filepath)
