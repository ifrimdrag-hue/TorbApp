"""Pricing module F0 one-off import + data-quality report.

Sources (commercial data folder, gitignored):
  FISIER_CONSOLIDAT_PRETURI.xlsx
    - sheet 'liste pret'              -> purchase prices into costuri_landing (fill
                                         missing rows for the target year; existing
                                         rows are compared, never overwritten)
    - per-client 'cod client X' cols  -> coduri_client_articol
    - sheet 'CODURI INTERNE CLIENTI'  -> coduri_client_articol (Metro dual codes)
    - sheet 'Cursuri'                 -> compared against rate_schimb (report only)
    - sheet 'CONDITII'                -> compared against cond_resolved (report only)
  RO1-010-26.xls / RO1-011-26.xls (Basilur order forms)
    - sheet 'Sheet2'                  -> produse_logistica (MC dims are cm in the
                                         file; converted to mm and validated against
                                         the CBM column)

Usage (from project root):
    .venv/bin/python etl/import_pricing_f0.py [--an 2026] [--dry-run]

Writes a full mismatch report next to the sources:
    Date pricinng&Logistica&Ofertare/rapoarte/f0_import_raport.txt
"""
import argparse
import os
import re
import sqlite3
import sys
import warnings

import openpyxl
import xlrd

warnings.filterwarnings("ignore")

DB = os.path.join("data", "torb.db")
DATA_DIR = "Date pricinng&Logistica&Ofertare"
START_DIR = os.path.join(DATA_DIR, "Fisiere start")
CONSOLIDAT = os.path.join(START_DIR, "FISIER_CONSOLIDAT_PRETURI.xlsx")
RO1_FILES = [os.path.join(START_DIR, f) for f in ("RO1-010-26.xls", "RO1-011-26.xls")]
REPORT_DIR = os.path.join(DATA_DIR, "rapoarte")

# per-client columns in 'liste pret': internal-code header, current invoicing
# price header (stripped), and the client-name search pattern in tranzactii
CLIENT_COLS = {
    "cod client AUCHAN":
        ("Pret facturare actual AUCHAN", "%AUCHAN%"),
    "Cod client METRO":
        ("Pret facturare actual Metro", "%METRO%"),
    "Cod Client Kaufland":
        ("Pret facturare actual kaufland", "%KAUFLAND%"),
    "Cod client EMAG RETAIL":
        ("Pret facturare actual EMAG RETAIL", "%EMAG RETAIL%"),
    "Cod client Sezamo":
        ("Pret facturare Sezamo", "%SEZAMO%"),
    "Cod client PROFI":  # not in tranzactii yet -> stays reported
        ("Pret facturare actual PROFI", "%PROFI ROM FOOD%"),
    "cod client SUPECO":
        ("Pret facturare actual SUPECO", "%SUPECO%"),
    "Cod client CARREFOUR":
        ("Pret facturare actual CARREFOUR", "%CARREFOUR%"),
}

report_lines = []


def rep(section, msg=""):
    line = f"[{section}] {msg}" if msg else section
    report_lines.append(line)


def resolve_sku(raw, sku_set):
    """Match a file article code against produse.sku (exact, +/- '-NN' suffix)."""
    if raw is None:
        return None
    c = str(raw).strip()
    if not c:
        return None
    if c in sku_set:
        return c
    if c + "-00" in sku_set:
        return c + "-00"
    base = re.sub(r"-\d+$", "", c)
    if base in sku_set:
        return base
    return None


def resolve_client(db, pattern):
    rows = db.execute(
        "SELECT DISTINCT cod_client, client FROM tranzactii "
        "WHERE upper(client) LIKE upper(?)", (pattern,)).fetchall()
    if len(rows) == 1:
        return rows[0]
    return None if not rows else ("AMBIGUU", [r[1] for r in rows])


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def import_consolidat(db, an, dry):
    wb = openpyxl.load_workbook(CONSOLIDAT, read_only=True, data_only=True)
    sku_set = {r[0] for r in db.execute("SELECT sku FROM produse")}

    ws = wb["liste pret"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr)}

    clients = {}
    for header, (_pret_hdr, pattern) in CLIENT_COLS.items():
        if header not in col:
            rep("CLIENTI", f"coloana lipsa in fisier: {header}")
            continue
        res = resolve_client(db, pattern)
        if res is None:
            rep("CLIENTI", f"client negasit in tranzactii pentru '{header}' ({pattern})")
        elif res[0] == "AMBIGUU":
            rep("CLIENTI", f"client ambiguu pentru '{header}': {res[1]}")
        else:
            clients[header] = res
            if not dry:
                db.execute(
                    "INSERT OR IGNORE INTO clienti_pricing(cod_client, nume_client) "
                    "VALUES (?, ?)", res)

    unmatched, land_new, land_diff, land_same, codes_ins = [], 0, 0, 0, 0
    pv_new, pv_diff = 0, 0
    landing = {r[0]: r for r in db.execute(
        "SELECT sku, moneda, pret_achizitie_valuta, transport_pct, taxa_vamala_pct "
        "FROM costuri_landing WHERE an = ?", (an,))}
    pv_exist = {(r[0], r[1]): r[2] for r in db.execute(
        "SELECT sku, cod_client, pret_vanzare_ron FROM preturi_vanzare "
        "WHERE an = ? AND cod_client IS NOT NULL", (an,))}
    cursuri_gama = {str(r[0]).strip().upper(): (str(r[1]).strip().upper(), num(r[2]))
                    for r in wb["Cursuri"].iter_rows(min_row=2, values_only=True)
                    if r[0] and r[1]}

    for r in rows[1:]:
        raw_cod = r[col["Cod articol"]]
        sku = resolve_sku(raw_cod, sku_set)
        if sku is None:
            if raw_cod is not None and str(raw_cod).strip():
                unmatched.append(f"{raw_cod}  |  {r[col['DENUMIRE']]}")
            continue

        # per-client internal codes + current invoicing prices
        for header, (cod_client, _n) in clients.items():
            val = r[col[header]]
            if val is not None and str(val).strip() not in ("", "0") and not dry:
                cur = db.execute(
                    "INSERT OR IGNORE INTO coduri_client_articol"
                    "(sku, cod_client, cod_intern, sursa) VALUES (?,?,?,?)",
                    (sku, cod_client, str(val).strip().rstrip(".0")
                     if isinstance(val, float) else str(val).strip(),
                     "FISIER_CONSOLIDAT"))
                codes_ins += cur.rowcount
            pret_hdr = CLIENT_COLS[header][0]
            pret_client = num(r[col[pret_hdr]]) if pret_hdr in col else None
            if pret_client:
                old = pv_exist.get((sku, cod_client))
                if old is None:
                    pv_new += 1
                    if not dry:
                        db.execute(
                            "INSERT OR IGNORE INTO preturi_vanzare"
                            "(an, sku, cod_client, pret_vanzare_ron, activ)"
                            " VALUES (?,?,?,?,1)", (an, sku, cod_client, pret_client))
                elif abs(old - pret_client) > 0.005:
                    pv_diff += 1
                    rep("PRET-CLIENT-DIFERIT",
                        f"{sku} / {cod_client}: DB {old} vs fisier {pret_client}")

        # purchase price -> costuri_landing (fill missing only)
        pret = num(r[col["Pret"]])
        if pret is None:
            continue
        gama = str(r[col["GAMA"]] or "").strip().upper()
        moneda, curs = cursuri_gama.get(gama, (None, None))
        curs_row = num(r[col["curs valuta import"]]) or curs
        transport = num(r[col["transport"]]) or 0.0
        taxe = num(r[col["TAXE"]]) or 0.0
        if sku in landing:
            old = landing[sku]
            if old[2] is not None and abs(old[2] - pret) > 0.005:
                land_diff += 1
                rep("PRET-DIFERIT",
                    f"{sku}: DB {old[2]} {old[1]} vs fisier {pret} {moneda or '?'}")
            else:
                land_same += 1
        else:
            if moneda is None or curs_row is None:
                rep("PRET-NOU-SKIP", f"{sku}: gama '{gama}' fara curs in sheet Cursuri")
                continue
            land_new += 1
            if not dry:
                ron = pret * curs_row
                db.execute(
                    "INSERT INTO costuri_landing(an, sku, moneda, pret_achizitie_valuta,"
                    " curs_ron, pret_achizitie_ron, transport_pct, taxa_vamala_pct,"
                    " alte_costuri_ron, landing_cost_ron) VALUES (?,?,?,?,?,?,?,?,0,?)",
                    (an, sku, moneda, pret, curs_row, ron, transport * 100,
                     taxe * 100, ron * (1 + transport) * (1 + taxe)))

    # Metro dual internal codes
    ws2 = wb["CODURI INTERNE CLIENTI"]
    metro = resolve_client(db, "%METRO%")
    for r in ws2.iter_rows(min_row=2, values_only=True):
        sku = resolve_sku(r[0], sku_set)
        if sku is None:
            if r[0]:
                unmatched.append(f"{r[0]}  |  {r[1]} (CODURI INTERNE)")
            continue
        if metro and metro[0] != "AMBIGUU" and str(r[5] or "").strip().upper() == "METRO":
            if not dry:
                db.execute(
                    "INSERT OR IGNORE INTO coduri_client_articol"
                    "(sku, cod_client, cod_intern, cod_intern2, sursa) VALUES (?,?,?,?,?)",
                    (sku, metro[0], str(r[3] or "").strip(), str(r[4] or "").strip(),
                     "FISIER_CONSOLIDAT"))

    # exchange rates: report only
    db_rate = {r[0]: r[1] for r in db.execute(
        "SELECT moneda, curs_ron FROM rate_schimb WHERE an = ?", (an,))}
    for gama, (mon, curs) in cursuri_gama.items():
        if mon in db_rate and curs is not None and abs(db_rate[mon] - curs) > 0.005:
            rep("CURS", f"{gama}/{mon}: fisier {curs} vs rate_schimb {db_rate[mon]}")
        elif mon not in db_rate:
            rep("CURS", f"{mon} (gama {gama}) lipseste din rate_schimb {an}")

    # CONDITII sheet vs cond_resolved: report only
    for r in wb["CONDITII"].iter_rows(min_row=2, values_only=True):
        nume, pct = str(r[0] or "").strip(), num(r[1])
        if not nume or pct is None:
            continue
        res = resolve_client(db, f"%{nume.split('-')[0]}%")
        if res is None or res[0] == "AMBIGUU":
            rep("CONDITII", f"'{nume}' ({pct:.2%}) - client nerezolvat, de mapat manual")
            continue
        eff = db.execute(
            "SELECT AVG(eff_pct) FROM cond_resolved WHERE an=? AND cod_client=?",
            (an, res[0])).fetchone()[0]
        if eff is None:
            rep("CONDITII", f"{nume} ({res[0]}): fisier {pct:.2%}, fara cond_resolved")
        elif abs(eff - pct * 100) > 0.5:
            rep("CONDITII", f"{nume} ({res[0]}): fisier {pct:.2%} vs cond_resolved {eff:.2f}%")

    wb.close()
    rep("SUMAR liste pret",
        f"landing: {land_new} noi, {land_same} identice, {land_diff} diferite; "
        f"coduri client inserate: {codes_ins}; preturi client: {pv_new} noi, "
        f"{pv_diff} diferite; SKU nepotrivite: {len(unmatched)}")
    if unmatched:
        rep("SKU NEPOTRIVITE (de creat in produse sau de mapat)")
        report_lines.extend("    " + u for u in unmatched)


def seed_conditii(db, an, dry):
    """Seed conditii_comerciale with per-client totals from the CONDITII sheet.

    One 'pct' row per resolved client (furnizor/categorie/sku = NULL = all),
    marked for later itemization (owner decision #4). Skips clients that
    already have any pct row for the year. cond_resolved is truncated so the
    app rebuilds it lazily at startup.
    """
    wb = openpyxl.load_workbook(CONSOLIDAT, read_only=True, data_only=True)
    ins = 0
    for r in wb["CONDITII"].iter_rows(min_row=2, values_only=True):
        nume, pct = str(r[0] or "").strip(), num(r[1])
        if not nume or pct is None:
            continue
        res = resolve_client(db, f"%{nume.split('-')[0]}%")
        if res is None or res[0] == "AMBIGUU":
            rep("SEED-CONDITII", f"'{nume}' ({pct:.2%}) nerezolvat - de introdus manual")
            continue
        exists = db.execute(
            "SELECT 1 FROM conditii_comerciale WHERE an=? AND cod_client=? "
            "AND tip_valoare='pct' LIMIT 1", (an, res[0])).fetchone()
        if exists:
            rep("SEED-CONDITII", f"{nume} ({res[0]}): are deja conditii pct - sarit")
            continue
        ins += 1
        if not dry:
            db.execute(
                "INSERT INTO conditii_comerciale(an, cod_client, furnizor,"
                " tip_valoare, periodicitate, valoare, descriere, data_creare)"
                " VALUES (?,?,NULL,'pct','anual',?,?,date('now'))",
                (an, res[0], round(pct * 100, 2),
                 "Total conditii client (import FISIER_CONSOLIDAT) - de defalcat"))
    wb.close()
    if ins and not dry:
        db.execute("DELETE FROM cond_resolved")  # lazy rebuild at app startup
    rep("SUMAR conditii", f"{ins} clienti seed-uiti cu % total conditii")


def import_logistica_basilur(db, dry):
    sku_set = {r[0] for r in db.execute("SELECT sku FROM produse")}
    produse = {r[0]: (r[1], r[2]) for r in db.execute(
        "SELECT sku, buc_cutie, gramaj FROM produse")}
    seen, ins, cbm_bad, buc_diff = set(), 0, 0, 0
    for path in RO1_FILES:
        if not os.path.exists(path):
            rep("LOGISTICA", f"fisier lipsa: {path}")
            continue
        ws = xlrd.open_workbook(path).sheet_by_name("Sheet2")
        hdr = [str(ws.cell_value(0, j)).strip() for j in range(ws.ncols)]
        col = {h: i for i, h in enumerate(hdr)}
        for i in range(1, ws.nrows):
            sku = resolve_sku(ws.cell_value(i, col["ItemCode"]), sku_set)
            if sku is None or sku in seen:
                continue
            seen.add(sku)
            g = lambda name: num(ws.cell_value(i, col[name]))  # noqa: E731
            units = g("No_of_Units")
            mc_net_g = g("Unit_Weight_g")           # net grams per master carton
            # MC_*_mm columns actually hold cm (validated against CBM below)
            lg, wd, ht = (v * 10 if v else None for v in
                          (g("MC_L_mm"), g("MC_W_mm"), g("MC_H_mm")))
            cbm = g("CBM")
            if lg and wd and ht and cbm and abs(lg * wd * ht / 1e9 - cbm) > cbm * 0.05:
                cbm_bad += 1
                rep("LOGISTICA-CBM", f"{sku}: dims {lg}x{wd}x{ht}mm -> "
                    f"{lg*wd*ht/1e9:.4f} m3 vs CBM {cbm}")
            unit_net = (mc_net_g / 1000 / units) if (mc_net_g and units) else None
            db_buc, db_gramaj = produse.get(sku, (None, None))
            if units and db_buc and int(units) != int(db_buc):
                buc_diff += 1
                rep("LOGISTICA-BUC", f"{sku}: fisier {int(units)}/bax vs "
                    f"produse.buc_cutie {db_buc}")
            if unit_net and db_gramaj and abs(unit_net * 1000 - db_gramaj) > 1:
                rep("LOGISTICA-GRAMAJ", f"{sku}: net/buc {unit_net*1000:.0f}g vs "
                    f"produse.gramaj {db_gramaj}")
            ins += 1
            if not dry:
                db.execute(
                    "INSERT INTO produse_logistica(sku, unit_net_kg, unit_gross_kg,"
                    " bax_l_mm, bax_w_mm, bax_h_mm, bax_gross_kg, bax_cbm, buc_bax,"
                    " sursa) VALUES (?,?,?,?,?,?,?,?,?,?) "
                    "ON CONFLICT(sku) DO UPDATE SET unit_net_kg=excluded.unit_net_kg,"
                    " unit_gross_kg=excluded.unit_gross_kg, bax_l_mm=excluded.bax_l_mm,"
                    " bax_w_mm=excluded.bax_w_mm, bax_h_mm=excluded.bax_h_mm,"
                    " bax_gross_kg=excluded.bax_gross_kg, bax_cbm=excluded.bax_cbm,"
                    " buc_bax=excluded.buc_bax, sursa=excluded.sursa,"
                    " updated_at=datetime('now','localtime')",
                    (sku, unit_net, g("Unit_Gross_Kg"), lg, wd, ht, g("MC_Gross_Kg"),
                     cbm, int(units) if units else None, os.path.basename(path)))
    rep("SUMAR logistica",
        f"{ins} SKU cu date logistice; CBM suspect: {cbm_bad}; buc/bax diferit: {buc_diff}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--an", type=int, default=2026)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--seed-conditii", action="store_true",
                    help="seed conditii_comerciale from the CONDITII sheet totals")
    args = ap.parse_args()

    if not os.path.exists(CONSOLIDAT):
        sys.exit(f"lipsa {CONSOLIDAT} - ruleaza din radacina proiectului")
    db = sqlite3.connect(DB)
    rep(f"F0 import - an {args.an}" + (" (DRY RUN)" if args.dry_run else ""))
    import_consolidat(db, args.an, args.dry_run)
    import_logistica_basilur(db, args.dry_run)
    if args.seed_conditii:
        seed_conditii(db, args.an, args.dry_run)
    if not args.dry_run:
        db.commit()
    db.close()

    os.makedirs(REPORT_DIR, exist_ok=True)
    out = os.path.join(REPORT_DIR, "f0_import_raport.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines) + "\n")
    summary = [ln for ln in report_lines if ln.startswith("[SUMAR")]
    print("\n".join(summary))
    print(f"raport complet: {out} ({len(report_lines)} linii)")


if __name__ == "__main__":
    main()
